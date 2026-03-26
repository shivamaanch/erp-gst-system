# modules/gst_reports.py
from flask import Blueprint, render_template, request, session, flash, redirect, url_for, send_file, jsonify
from flask_login import login_required
from extensions import db
from models import Bill, Party, BillItem, Item, Gstr2bRecord
from sqlalchemy import func
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json

gst_reports_bp = Blueprint("gst_reports", __name__)

# ═══════════════════════════════════════════════════════
#  GSTR-1 - OUTWARD SUPPLIES (SALES)
# ═══════════════════════════════════════════════════════

@gst_reports_bp.route("/gst/gstr1")
@login_required
def gstr1():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    export = request.args.get("export", "")
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get all sales invoices for the period
    bills = Bill.query.filter(
        Bill.company_id == cid,
        Bill.bill_type == "Sales",
        Bill.is_cancelled == False,
        Bill.bill_date >= start_date,
        Bill.bill_date <= end_date
    ).order_by(Bill.bill_date).all()
    
    # Categorize invoices
    b2b_invoices = []  # Business to Business (with GSTIN)
    b2c_large = []     # B2C Large (>2.5 lakh)
    b2c_small = []     # B2C Small (<2.5 lakh)
    
    total_taxable = 0
    total_cgst = 0
    total_sgst = 0
    total_igst = 0
    total_invoice_value = 0
    
    for bill in bills:
        party = Party.query.get(bill.party_id)
        invoice_data = {
            "bill_no": bill.bill_no,
            "bill_date": bill.bill_date,
            "party_name": party.name if party else "Unknown",
            "gstin": party.gstin if party else None,
            "taxable": float(bill.taxable_amount or 0),
            "cgst": float(bill.cgst or 0),
            "sgst": float(bill.sgst or 0),
            "igst": float(bill.igst or 0),
            "total": float(bill.total_amount or 0),
            "state_code": party.state_code if party else None
        }
        
        total_taxable += invoice_data["taxable"]
        total_cgst += invoice_data["cgst"]
        total_sgst += invoice_data["sgst"]
        total_igst += invoice_data["igst"]
        total_invoice_value += invoice_data["total"]
        
        if party and party.gstin:
            # B2B Invoice
            b2b_invoices.append(invoice_data)
        elif invoice_data["total"] > 250000:
            # B2C Large
            b2c_large.append(invoice_data)
        else:
            # B2C Small
            b2c_small.append(invoice_data)
    
    # Group B2C Small by state and rate
    b2c_small_summary = {}
    for inv in b2c_small:
        state = inv["state_code"] or "Unknown"
        rate = 0
        if inv["cgst"] > 0:
            rate = round((inv["cgst"] / inv["taxable"]) * 100, 2) if inv["taxable"] > 0 else 0
        elif inv["igst"] > 0:
            rate = round((inv["igst"] / inv["taxable"]) * 100, 2) if inv["taxable"] > 0 else 0
        
        key = f"{state}_{rate}"
        if key not in b2c_small_summary:
            b2c_small_summary[key] = {
                "state": state,
                "rate": rate,
                "taxable": 0,
                "cgst": 0,
                "sgst": 0,
                "igst": 0,
                "count": 0
            }
        b2c_small_summary[key]["taxable"] += inv["taxable"]
        b2c_small_summary[key]["cgst"] += inv["cgst"]
        b2c_small_summary[key]["sgst"] += inv["sgst"]
        b2c_small_summary[key]["igst"] += inv["igst"]
        b2c_small_summary[key]["count"] += 1
    
    if export == "excel":
        return export_gstr1_excel(b2b_invoices, b2c_large, list(b2c_small_summary.values()), 
                                   period, total_taxable, total_cgst, total_sgst, total_igst)
    
    return render_template("gst_reports/gstr1.html",
                         period=period,
                         b2b_invoices=b2b_invoices,
                         b2c_large=b2c_large,
                         b2c_small=list(b2c_small_summary.values()),
                         total_taxable=round(total_taxable, 2),
                         total_cgst=round(total_cgst, 2),
                         total_sgst=round(total_sgst, 2),
                         total_igst=round(total_igst, 2),
                         total_invoice_value=round(total_invoice_value, 2),
                         b2b_count=len(b2b_invoices),
                         b2c_large_count=len(b2c_large))

# ═══════════════════════════════════════════════════════
#  GSTR-2B - INWARD SUPPLIES (PURCHASES) - IMPORT & RECONCILE
# ═══════════════════════════════════════════════════════

@gst_reports_bp.route("/gst/gstr2b", methods=["GET", "POST"])
@login_required
def gstr2b():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    if request.method == "POST":
        # Handle file upload (JSON or Excel)
        if "file" in request.files:
            file = request.files["file"]
            if file.filename.endswith(".json"):
                return import_gstr2b_json(file, cid, fy, period)
            elif file.filename.endswith((".xlsx", ".xls")):
                return import_gstr2b_excel(file, cid, fy, period)
    
    # Fetch imported 2B records
    records = Gstr2bRecord.query.filter_by(
        company_id=cid,
        fin_year=fy,
        period=period
    ).order_by(Gstr2bRecord.invoice_date).all()
    
    total_taxable = sum(float(r.taxable_value or 0) for r in records)
    total_igst = sum(float(r.igst or 0) for r in records)
    total_cgst = sum(float(r.cgst or 0) for r in records)
    total_sgst = sum(float(r.sgst or 0) for r in records)
    
    return render_template("gst_reports/gstr2b.html",
                         period=period,
                         records=records,
                         total_taxable=round(total_taxable, 2),
                         total_igst=round(total_igst, 2),
                         total_cgst=round(total_cgst, 2),
                         total_sgst=round(total_sgst, 2),
                         record_count=len(records))

def import_gstr2b_json(file, cid, fy, period):
    """Import GSTR-2B from GSTN portal JSON"""
    try:
        data = json.load(file)
        # Clear existing records for this period
        Gstr2bRecord.query.filter_by(company_id=cid, fin_year=fy, period=period).delete()
        
        imported = 0
        # Parse JSON structure (GSTN format)
        if "data" in data and "docdata" in data["data"]:
            for doc in data["data"]["docdata"].get("b2b", []):
                for inv in doc.get("inv", []):
                    record = Gstr2bRecord(
                        company_id=cid,
                        fin_year=fy,
                        period=period,
                        supplier_gstin=doc.get("ctin"),
                        supplier_name=doc.get("trdnm", "Unknown"),
                        invoice_no=inv.get("inum"),
                        invoice_date=datetime.strptime(inv.get("idt"), "%d-%m-%Y").date() if inv.get("idt") else None,
                        invoice_type="B2B",
                        taxable_value=float(inv.get("val", 0)),
                        igst=float(inv.get("igst", 0)),
                        cgst=float(inv.get("cgst", 0)),
                        sgst=float(inv.get("sgst", 0)),
                        itc_available=True,
                        status="pending"
                    )
                    db.session.add(record)
                    imported += 1
        
        db.session.commit()
        flash(f"✅ Successfully imported {imported} records from GSTR-2B JSON", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error importing JSON: {str(e)}", "danger")
    
    return redirect(url_for("gst_reports.gstr2b", period=period))

def import_gstr2b_excel(file, cid, fy, period):
    """Import GSTR-2B from Excel file"""
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Clear existing records
        Gstr2bRecord.query.filter_by(company_id=cid, fin_year=fy, period=period).delete()
        
        imported = 0
        # Expected columns: GSTIN, Supplier Name, Invoice No, Invoice Date, Taxable Value, IGST, CGST, SGST
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            record = Gstr2bRecord(
                company_id=cid,
                fin_year=fy,
                period=period,
                supplier_gstin=str(row[0]) if row[0] else None,
                supplier_name=str(row[1]) if row[1] else "Unknown",
                invoice_no=str(row[2]) if row[2] else None,
                invoice_date=row[3] if isinstance(row[3], date) else None,
                invoice_type="B2B",
                taxable_value=float(row[4]) if row[4] else 0,
                igst=float(row[5]) if row[5] else 0,
                cgst=float(row[6]) if row[6] else 0,
                sgst=float(row[7]) if row[7] else 0,
                itc_available=True,
                status="pending"
            )
            db.session.add(record)
            imported += 1
        
        db.session.commit()
        flash(f"✅ Successfully imported {imported} records from Excel", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error importing Excel: {str(e)}", "danger")
    
    return redirect(url_for("gst_reports.gstr2b", period=period))

# ═══════════════════════════════════════════════════════
#  GSTR-3B - MONTHLY SUMMARY RETURN
# ═══════════════════════════════════════════════════════

@gst_reports_bp.route("/gst/gstr3b")
@login_required
def gstr3b():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # OUTWARD SUPPLIES (Sales)
    sales = Bill.query.filter(
        Bill.company_id == cid,
        Bill.bill_type == "Sales",
        Bill.is_cancelled == False,
        Bill.bill_date >= start_date,
        Bill.bill_date <= end_date
    ).all()
    
    outward_taxable = sum(float(b.taxable_amount or 0) for b in sales)
    outward_igst = sum(float(b.igst or 0) for b in sales)
    outward_cgst = sum(float(b.cgst or 0) for b in sales)
    outward_sgst = sum(float(b.sgst or 0) for b in sales)
    
    # INWARD SUPPLIES (Purchases)
    purchases = Bill.query.filter(
        Bill.company_id == cid,
        Bill.bill_type == "Purchase",
        Bill.is_cancelled == False,
        Bill.bill_date >= start_date,
        Bill.bill_date <= end_date
    ).all()
    
    inward_taxable = sum(float(b.taxable_amount or 0) for b in purchases)
    inward_igst = sum(float(b.igst or 0) for b in purchases)
    inward_cgst = sum(float(b.cgst or 0) for b in purchases)
    inward_sgst = sum(float(b.sgst or 0) for b in purchases)
    
    # ITC Available (from purchases)
    itc_igst = inward_igst
    itc_cgst = inward_cgst
    itc_sgst = inward_sgst
    
    # Net Tax Liability
    net_igst = max(0, outward_igst - itc_igst)
    net_cgst = max(0, outward_cgst - itc_cgst)
    net_sgst = max(0, outward_sgst - itc_sgst)
    total_tax_liability = net_igst + net_cgst + net_sgst
    
    return render_template("gst_reports/gstr3b.html",
                         period=period,
                         outward_taxable=round(outward_taxable, 2),
                         outward_igst=round(outward_igst, 2),
                         outward_cgst=round(outward_cgst, 2),
                         outward_sgst=round(outward_sgst, 2),
                         inward_taxable=round(inward_taxable, 2),
                         inward_igst=round(inward_igst, 2),
                         inward_cgst=round(inward_cgst, 2),
                         inward_sgst=round(inward_sgst, 2),
                         itc_igst=round(itc_igst, 2),
                         itc_cgst=round(itc_cgst, 2),
                         itc_sgst=round(itc_sgst, 2),
                         net_igst=round(net_igst, 2),
                         net_cgst=round(net_cgst, 2),
                         net_sgst=round(net_sgst, 2),
                         total_tax_liability=round(total_tax_liability, 2))

# ═══════════════════════════════════════════════════════
#  GST RECONCILIATION - 2B vs BOOKS
# ═══════════════════════════════════════════════════════

@gst_reports_bp.route("/gst/reconcile")
@login_required
def reconcile():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get GSTR-2B records
    gstr2b_records = Gstr2bRecord.query.filter_by(
        company_id=cid,
        fin_year=fy,
        period=period
    ).all()
    
    # Get Purchase Bills from books
    purchase_bills = Bill.query.filter(
        Bill.company_id == cid,
        Bill.bill_type == "Purchase",
        Bill.is_cancelled == False,
        Bill.bill_date >= start_date,
        Bill.bill_date <= end_date
    ).all()
    
    # Create reconciliation data
    matched = []
    missing_in_books = []
    missing_in_2b = []
    mismatched = []
    
    # Create lookup dictionaries
    gstr2b_lookup = {f"{r.supplier_gstin}_{r.invoice_no}": r for r in gstr2b_records}
    books_lookup = {}
    
    for bill in purchase_bills:
        party = Party.query.get(bill.party_id)
        if party and party.gstin:
            key = f"{party.gstin}_{bill.bill_no}"
            books_lookup[key] = {"bill": bill, "party": party}
    
    # Find matches and mismatches
    for key, gstr2b_rec in gstr2b_lookup.items():
        if key in books_lookup:
            book_data = books_lookup[key]
            bill = book_data["bill"]
            
            # Check if amounts match
            book_total = float(bill.taxable_amount or 0)
            gstr2b_total = float(gstr2b_rec.taxable_value or 0)
            diff = abs(book_total - gstr2b_total)
            
            if diff < 0.01:  # Matched
                matched.append({
                    "gstin": gstr2b_rec.supplier_gstin,
                    "supplier": gstr2b_rec.supplier_name,
                    "invoice_no": gstr2b_rec.invoice_no,
                    "invoice_date": gstr2b_rec.invoice_date,
                    "taxable_2b": gstr2b_total,
                    "taxable_books": book_total,
                    "diff": 0,
                    "status": "Matched"
                })
            else:  # Mismatched
                mismatched.append({
                    "gstin": gstr2b_rec.supplier_gstin,
                    "supplier": gstr2b_rec.supplier_name,
                    "invoice_no": gstr2b_rec.invoice_no,
                    "invoice_date": gstr2b_rec.invoice_date,
                    "taxable_2b": gstr2b_total,
                    "taxable_books": book_total,
                    "diff": round(diff, 2),
                    "status": "Amount Mismatch"
                })
        else:
            # Missing in books
            missing_in_books.append({
                "gstin": gstr2b_rec.supplier_gstin,
                "supplier": gstr2b_rec.supplier_name,
                "invoice_no": gstr2b_rec.invoice_no,
                "invoice_date": gstr2b_rec.invoice_date,
                "taxable_2b": float(gstr2b_rec.taxable_value or 0),
                "status": "Missing in Books"
            })
    
    # Find invoices in books but not in 2B
    for key, book_data in books_lookup.items():
        if key not in gstr2b_lookup:
            bill = book_data["bill"]
            party = book_data["party"]
            missing_in_2b.append({
                "gstin": party.gstin,
                "supplier": party.name,
                "invoice_no": bill.bill_no,
                "invoice_date": bill.bill_date,
                "taxable_books": float(bill.taxable_amount or 0),
                "status": "Missing in GSTR-2B"
            })
    
    return render_template("gst_reports/reconcile.html",
                         period=period,
                         matched=matched,
                         missing_in_books=missing_in_books,
                         missing_in_2b=missing_in_2b,
                         mismatched=mismatched,
                         matched_count=len(matched),
                         missing_books_count=len(missing_in_books),
                         missing_2b_count=len(missing_in_2b),
                         mismatch_count=len(mismatched))

# ═══════════════════════════════════════════════════════
#  EXCEL EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════

def export_gstr1_excel(b2b, b2c_large, b2c_small, period, total_taxable, total_cgst, total_sgst, total_igst):
    """Export GSTR-1 to Excel in GSTN format"""
    wb = openpyxl.Workbook()
    
    # B2B Sheet
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    headers = ["GSTIN", "Receiver Name", "Invoice No", "Invoice Date", "Taxable Value", "CGST", "SGST", "IGST", "Total"]
    ws_b2b.append(headers)
    
    for inv in b2b:
        ws_b2b.append([
            inv["gstin"], inv["party_name"], inv["bill_no"], inv["bill_date"].strftime("%d-%m-%Y"),
            inv["taxable"], inv["cgst"], inv["sgst"], inv["igst"], inv["total"]
        ])
    
    # B2C Large Sheet
    ws_b2c_large = wb.create_sheet("B2C Large")
    ws_b2c_large.append(["Invoice No", "Invoice Date", "Taxable Value", "CGST", "SGST", "IGST", "Total"])
    for inv in b2c_large:
        ws_b2c_large.append([
            inv["bill_no"], inv["bill_date"].strftime("%d-%m-%Y"),
            inv["taxable"], inv["cgst"], inv["sgst"], inv["igst"], inv["total"]
        ])
    
    # B2C Small Sheet
    ws_b2c_small = wb.create_sheet("B2C Small")
    ws_b2c_small.append(["State", "Rate", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Count"])
    for summary in b2c_small:
        ws_b2c_small.append([
            summary["state"], summary["rate"], summary["taxable"],
            summary["cgst"], summary["sgst"], summary["igst"], summary["count"]
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name=f"GSTR1_{period}.xlsx", as_attachment=True,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
