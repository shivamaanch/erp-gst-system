# modules/tds_tcs.py
from flask import Blueprint, render_template, request, session, flash, redirect, url_for, send_file
from flask_login import login_required
from extensions import db
from models import Bill, Party, Company
from datetime import datetime, date
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

tds_tcs_bp = Blueprint("tds_tcs", __name__)

# TDS Rates based on sections
TDS_RATES = {
    "194C": {"rate": 0.02, "threshold": 10000, "section": "Contract"},
    "194J": {"rate": 0.10, "threshold": 30000, "section": "Professional Fees"},
    "194I": {"rate": 0.10, "threshold": 5000, "section": "Interest"},
    "194A": {"rate": 0.10, "threshold": 15000, "section": "Rent"},
    "192A": {"rate": 0.02, "threshold": 50000, "section": "Commission"},
    "default": {"rate": 0.01, "threshold": 50000, "section": "Others"}
}

# TCS Rates
TCS_RATES = {
    "206C": {"rate": 0.01, "section": "Sale of Motor Vehicle"},
    "206CII": {"rate": 0.05, "section": "Scrap"},
    "206CI": {"rate": 0.01, "section": "Alcohol"},
    "default": {"rate": 0.01, "section": "Others"}
}

@tds_tcs_bp.route("/tds")
@login_required
def tds_index():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get purchase bills with TDS
    bills = db.session.query(Bill, Party).join(Party, Bill.party_id==Party.id).filter(
        Bill.company_id==cid, Bill.fin_year==fy,
        Bill.bill_type=="Purchase", Bill.is_cancelled==False,
        Bill.bill_date >= start_date, Bill.bill_date <= end_date,
        Bill.tds_amount > 0
    ).order_by(Bill.bill_date.desc()).all()
    
    # Calculate totals
    total_taxable = sum(float(bill.taxable_amount or 0) for bill, party in bills)
    total_tds = sum(float(bill.tds_amount or 0) for bill, party in bills)
    
    return render_template("tds_tcs/tds_index.html", 
                         bills=bills, period=period,
                         total_taxable=round(total_taxable, 2),
                         total_tds=round(total_tds, 2),
                         bill_count=len(bills))

@tds_tcs_bp.route("/tcs")
@login_required
def tcs_index():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get sales bills with TCS
    bills = db.session.query(Bill, Party).join(Party, Bill.party_id==Party.id).filter(
        Bill.company_id==cid, Bill.fin_year==fy,
        Bill.bill_type=="Sales", Bill.is_cancelled==False,
        Bill.bill_date >= start_date, Bill.bill_date <= end_date,
        Bill.tcs_amount > 0
    ).order_by(Bill.bill_date.desc()).all()
    
    # Calculate totals
    total_taxable = sum(float(bill.taxable_amount or 0) for bill, party in bills)
    total_tcs = sum(float(bill.tcs_amount or 0) for bill, party in bills)
    
    return render_template("tds_tcs/tcs_index.html", 
                         bills=bills, period=period,
                         total_taxable=round(total_taxable, 2),
                         total_tcs=round(total_tcs, 2),
                         bill_count=len(bills))

@tds_tcs_bp.route("/tds/calculate/<int:bill_id>", methods=["GET", "POST"])
@login_required
def calculate_tds(bill_id):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    
    if request.method == "POST":
        section = request.form.get("tds_section", "default")
        tds_info = TDS_RATES.get(section, TDS_RATES["default"])
        
        # Calculate TDS
        taxable_amount = float(bill.taxable_amount or 0)
        if taxable_amount >= tds_info["threshold"]:
            tds_rate = tds_info["rate"]
            tds_amount = round(taxable_amount * tds_rate, 2)
        else:
            tds_rate = 0
            tds_amount = 0
        
        bill.tds_rate = tds_rate
        bill.tds_amount = tds_amount
        bill.total_amount = taxable_amount + float(bill.cgst or 0) + float(bill.sgst or 0) + float(bill.igst or 0) - tds_amount
        db.session.commit()
        
        flash(f"TDS calculated: ₹{tds_amount} ({tds_rate*100}%)", "success")
        return redirect(url_for("tds_tcs.tds_index"))
    
    # Get current TDS info
    current_section = "default"
    for section, info in TDS_RATES.items():
        if bill.tds_rate == info["rate"]:
            current_section = section
            break
    
    return render_template("tds_tcs/calculate_tds.html", 
                         bill=bill, party=party,
                         tds_sections=TDS_RATES,
                         current_section=current_section)

@tds_tcs_bp.route("/tcs/calculate/<int:bill_id>", methods=["GET", "POST"])
@login_required
def calculate_tcs(bill_id):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    
    if request.method == "POST":
        section = request.form.get("tcs_section", "default")
        tcs_info = TCS_RATES.get(section, TCS_RATES["default"])
        
        # Calculate TCS
        taxable_amount = float(bill.taxable_amount or 0)
        tcs_rate = tcs_info["rate"]
        tcs_amount = round(taxable_amount * tcs_rate, 2)
        
        bill.tcs_rate = tcs_rate
        bill.tcs_amount = tcs_amount
        bill.total_amount = taxable_amount + float(bill.cgst or 0) + float(bill.sgst or 0) + float(bill.igst or 0) + tcs_amount
        db.session.commit()
        
        flash(f"TCS calculated: ₹{tcs_amount} ({tcs_rate*100}%)", "success")
        return redirect(url_for("tds_tcs.tcs_index"))
    
    # Get current TCS info
    current_section = "default"
    for section, info in TCS_RATES.items():
        if bill.tcs_rate == info["rate"]:
            current_section = section
            break
    
    return render_template("tds_tcs/calculate_tcs.html", 
                         bill=bill, party=party,
                         tcs_sections=TCS_RATES,
                         current_section=current_section)

@tds_tcs_bp.route("/tds/certificate/<int:bill_id>")
@login_required
def tds_certificate(bill_id):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    company = Company.query.get(cid)
    
    return render_template("tds_tcs/tds_certificate.html", 
                         bill=bill, party=party, company=company)

@tds_tcs_bp.route("/tcs/certificate/<int:bill_id>")
@login_required
def tcs_certificate(bill_id):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    company = Company.query.get(cid)
    
    return render_template("tds_tcs/tcs_certificate.html", 
                         bill=bill, party=party, company=company)

@tds_tcs_bp.route("/tds/export")
@login_required
def export_tds():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get data
    bills = db.session.query(Bill, Party).join(Party, Bill.party_id==Party.id).filter(
        Bill.company_id==cid, Bill.fin_year==fy,
        Bill.bill_type=="Purchase", Bill.is_cancelled==False,
        Bill.bill_date >= start_date, Bill.bill_date <= end_date,
        Bill.tds_amount > 0
    ).order_by(Bill.bill_date).all()
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDS Details"
    
    # Headers
    headers = ["Bill No", "Date", "Party", "PAN", "GSTIN", "Taxable Amount", "TDS Rate", "TDS Amount", "Total Amount"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Data
    for bill, party in bills:
        ws.append([
            bill.bill_no,
            bill.bill_date.strftime("%d-%m-%Y"),
            party.name,
            party.pan or "Not Available",
            party.gstin or "Not Available",
            float(bill.taxable_amount or 0),
            float(bill.tds_rate or 0) * 100,
            float(bill.tds_amount or 0),
            float(bill.total_amount or 0)
        ])
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name=f"TDS_{period}.xlsx", as_attachment=True,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@tds_tcs_bp.route("/tcs/export")
@login_required
def export_tcs():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))
    
    # Parse period
    try:
        month, year = period.split("-")
        start_date = date(int(year), int(month), 1)
        if int(month) == 12:
            end_date = date(int(year)+1, 1, 1)
        else:
            end_date = date(int(year), int(month)+1, 1)
        from datetime import timedelta
        end_date = end_date - timedelta(days=1)
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get data
    bills = db.session.query(Bill, Party).join(Party, Bill.party_id==Party.id).filter(
        Bill.company_id==cid, Bill.fin_year==fy,
        Bill.bill_type=="Sales", Bill.is_cancelled==False,
        Bill.bill_date >= start_date, Bill.bill_date <= end_date,
        Bill.tcs_amount > 0
    ).order_by(Bill.bill_date).all()
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TCS Details"
    
    # Headers
    headers = ["Bill No", "Date", "Customer", "PAN", "GSTIN", "Taxable Amount", "TCS Rate", "TCS Amount", "Total Amount"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Data
    for bill, party in bills:
        ws.append([
            bill.bill_no,
            bill.bill_date.strftime("%d-%m-%Y"),
            party.name,
            party.pan or "Not Available",
            party.gstin or "Not Available",
            float(bill.taxable_amount or 0),
            float(bill.tcs_rate or 0) * 100,
            float(bill.tcs_amount or 0),
            float(bill.total_amount or 0)
        ])
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name=f"TCS_{period}.xlsx", as_attachment=True,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
