from flask import Blueprint, render_template, request, session, flash, redirect, url_for, Response
from flask_login import login_required, current_user
from extensions import db
from models import CashBook, Company, FinancialYear, Account, JournalHeader, JournalLine, BankAccount, Party
from datetime import date, datetime
from sqlalchemy import text
from utils.filters import fy_dates

cash_book_bp = Blueprint("cash_book", __name__)

def next_voucher_no(company_id, fin_year):
    """Generate next voucher number for cash book"""
    last = CashBook.query.filter_by(company_id=company_id, fin_year=fin_year).order_by(CashBook.id.desc()).first()
    if last and last.voucher_no:
        try:
            num = int(last.voucher_no.split('-')[-1]) + 1
        except:
            num = 1
    else:
        num = 1
    return f"CB-{fin_year}-{num:04d}"

def next_journal_voucher_no(company_id, fin_year):
    """Generate next voucher number for journal"""
    from models import JournalHeader
    last = JournalHeader.query.filter_by(company_id=company_id, fin_year=fin_year).order_by(JournalHeader.id.desc()).first()
    if last and last.voucher_no:
        try:
            num = int(last.voucher_no.split('-')[-1]) + 1
        except:
            num = 1
    else:
        num = 1
    return f"JNL-{fin_year}-{num:04d}"

@cash_book_bp.route("/cash-book")
@login_required
def index():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    # Get filters
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    narration_filter = request.args.get("narration", "")
    search_term = request.args.get("search", "")
    
    # Get default cash account (first account with 'Cash' in name or ID 1)
    default_cash = Account.query.filter_by(company_id=cid, is_active=True).filter(
        Account.name.ilike('%cash%')
    ).first()
    if not default_cash and Account.query.filter_by(company_id=cid, is_active=True).first():
        default_cash = Account.query.filter_by(company_id=cid, is_active=True).first()
    
    # Default date to last used or today
    default_date = session.get("last_txn_date") or date.today().isoformat()
    
    # Query cash book entries
    query = CashBook.query.filter_by(company_id=cid, fin_year=fy)
    
    # Apply filters
    if from_date:
        try:
            query = query.filter(CashBook.transaction_date >= datetime.strptime(from_date, "%Y-%m-%d").date())
        except:
            pass
    if to_date:
        try:
            query = query.filter(CashBook.transaction_date <= datetime.strptime(to_date, "%Y-%m-%d").date())
        except:
            pass
    if narration_filter:
        query = query.filter(CashBook.narration.ilike(f"%{narration_filter}%"))
    if search_term:
        query = query.filter(
            CashBook.narration.ilike(f"%{search_term}%") |
            CashBook.party_name.ilike(f"%{search_term}%") |
            CashBook.voucher_no.ilike(f"%{search_term}%")
        )
    
    cash_entries = query.order_by(CashBook.transaction_date.asc(), CashBook.id.asc()).all()
    
    # Calculate opening balance as of from_date (moved up)
    opening_balance = 0.0
    try:
        if from_date:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            opening_entries = CashBook.query.filter(
                CashBook.company_id == cid,
                CashBook.fin_year == fy,
                CashBook.transaction_date < from_date_obj
            ).all()
        else:
            # If no from_date, get all entries before the first entry in the current period
            if cash_entries:
                first_entry_date = cash_entries[0].transaction_date
                opening_entries = CashBook.query.filter(
                    CashBook.company_id == cid,
                    CashBook.fin_year == fy,
                    CashBook.transaction_date < first_entry_date
                ).all()
            else:
                opening_entries = []
        
        opening_balance = sum(entry.amount for entry in opening_entries if entry.transaction_type == "Receipt") - \
                      sum(entry.amount for entry in opening_entries if entry.transaction_type == "Payment")
    except:
        pass
    
    # Calculate cumulative balances for each entry
    entries_with_balances = []
    running_balance = opening_balance
    
    for entry in cash_entries:
        # Calculate balance change for this entry
        if entry.transaction_type == 'Receipt':
            running_balance += float(entry.amount)
        elif entry.transaction_type == 'Payment':
            running_balance -= float(entry.amount)
        
        # Add entry with its cumulative balance
        entry_with_balance = {
            'entry': entry,
            'cumulative_balance': running_balance,
            'cash_received': float(entry.amount) if entry.transaction_type == 'Receipt' else 0.0,
            'cash_paid': float(entry.amount) if entry.transaction_type == 'Payment' else 0.0
        }
        entries_with_balances.append(entry_with_balance)
    
    # Calculate totals
    total_receipts = sum(entry.amount for entry in cash_entries if entry.transaction_type == "Receipt")
    total_payments = sum(entry.amount for entry in cash_entries if entry.transaction_type == "Payment")
    balance = total_receipts - total_payments
    
    # Calculate cash negative during period
    cash_negative = max(0, total_payments - total_receipts)
    
    # Calculate maximum negative cash and when it started
    max_negative_cash = 0.0
    cash_negative_start_date = None
    running_balance = opening_balance
    
    # Sort entries by date to track cash flow
    sorted_entries = sorted(cash_entries, key=lambda x: x.transaction_date)
    
    for entry in sorted_entries:
        if entry.transaction_type == "Receipt":
            running_balance += float(entry.amount)
        else:  # Payment
            running_balance -= float(entry.amount)
        
        if running_balance < 0:
            if cash_negative_start_date is None:
                cash_negative_start_date = entry.transaction_date
            
            if abs(running_balance) > max_negative_cash:
                max_negative_cash = abs(running_balance)
    
    # Get client mobile numbers for WhatsApp sharing
    client_mobiles = {}
    for entry in cash_entries:
        if entry.account_id:
            account = Account.query.get(entry.account_id)
            if account and hasattr(account, 'phone') and account.phone:
                client_mobiles[entry.id] = account.phone
    
    return render_template("cash_book/index.html",
                         entries_with_balances=entries_with_balances,
                         total_receipts=total_receipts,
                         total_payments=total_payments,
                         balance=balance,
                         opening_balance=opening_balance,
                         cash_negative=cash_negative,
                         max_negative_cash=max_negative_cash,
                         cash_negative_start_date=cash_negative_start_date,
                         from_date=from_date or "",
                         to_date=to_date or "",
                         narration_filter=narration_filter,
                         search_term=search_term,
                         default_cash=default_cash,
                         client_mobiles=client_mobiles,
                         fy=fy,
                         default_date=default_date)

@cash_book_bp.route("/cash_book/export")
@login_required
def export_cash_book():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    export_type = request.args.get("export", "pdf")
    
    # Get same filters as index
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    narration_filter = request.args.get("narration", "")
    search_term = request.args.get("search", "")
    
    # Query cash book entries
    query = CashBook.query.filter_by(company_id=cid, fin_year=fy)
    
    # Apply filters
    if from_date:
        try:
            query = query.filter(CashBook.transaction_date >= datetime.strptime(from_date, "%Y-%m-%d").date())
        except:
            pass
    if to_date:
        try:
            query = query.filter(CashBook.transaction_date <= datetime.strptime(to_date, "%Y-%m-%d").date())
        except:
            pass
    if narration_filter:
        query = query.filter(CashBook.narration.ilike(f"%{narration_filter}%"))
    if search_term:
        query = query.filter(
            CashBook.narration.ilike(f"%{search_term}%") |
            CashBook.party_name.ilike(f"%{search_term}%") |
            CashBook.voucher_no.ilike(f"%{search_term}%")
        )
    
    cash_entries = query.order_by(CashBook.transaction_date.asc(), CashBook.id.asc()).all()
    
    # Calculate totals
    total_receipts = sum(entry.amount for entry in cash_entries if entry.transaction_type == "Receipt")
    total_payments = sum(entry.amount for entry in cash_entries if entry.transaction_type == "Payment")
    balance = total_receipts - total_payments
    
    # Calculate opening balance
    opening_balance = 0.0
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            opening_entries = CashBook.query.filter(
                CashBook.company_id == cid,
                CashBook.fin_year == fy,
                CashBook.transaction_date < from_date_obj
            ).all()
            opening_balance = sum(entry.amount for entry in opening_entries if entry.transaction_type == "Receipt") - \
                          sum(entry.amount for entry in opening_entries if entry.transaction_type == "Payment")
        except:
            pass
    
    # Calculate cash negative during period
    cash_negative = max(0, total_payments - total_receipts)
    
    if export_type == "pdf":
        return export_cash_book_pdf(cash_entries, total_receipts, total_payments, balance, 
                                    opening_balance, cash_negative, from_date, to_date, fy)
    elif export_type == "excel":
        return export_cash_book_excel(cash_entries, total_receipts, total_payments, balance,
                                       opening_balance, cash_negative, from_date, to_date, fy)
    else:
        return redirect(url_for("cash_book.index"))

def export_cash_book_pdf(cash_entries, total_receipts, total_payments, balance, 
                            opening_balance, cash_negative, from_date, to_date, fy):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Frame, PageTemplate
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.units import inch, cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    from flask import Response
    from datetime import datetime
    
    # Create buffer first
    buffer = BytesIO()
    
    # Create PDF document with A4 portrait
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Add custom styles
    styles.add(ParagraphStyle(
        'CompanyTitle', 
        parent=styles['Heading1'],
        fontSize=18, 
        textColor=colors.black, 
        alignment=TA_CENTER,
        spaceAfter=0.2*cm
    ))
    
    styles.add(ParagraphStyle(
        'ReportTitle', 
        parent=styles['Heading2'],
        fontSize=14, 
        textColor=colors.black, 
        alignment=TA_CENTER,
        spaceAfter=0.3*cm,
        spaceBefore=0.1*cm
    ))
    
    styles.add(ParagraphStyle(
        'PeriodInfo', 
        parent=styles['Normal'],
        fontSize=10, 
        textColor=colors.black, 
        alignment=TA_CENTER,
        spaceAfter=0.5*cm
    ))
    
    styles.add(ParagraphStyle(
        'SummaryLabel', 
        parent=styles['Normal'],
        fontSize=9, 
        textColor=colors.black, 
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    ))
    
    styles.add(ParagraphStyle(
        'SummaryValue', 
        parent=styles['Normal'],
        fontSize=9, 
        textColor=colors.black, 
        alignment=TA_RIGHT,
        fontName='Helvetica-Bold'
    ))
    
    # Add content
    story = []
    
    # Company Header (would need to get from session/database)
    story.append(Paragraph("SHIVA ICE & MILK PRODUCTS", styles['CompanyTitle']))
    story.append(Paragraph("Cash Book Statement", styles['ReportTitle']))
    
    # Period Information
    period_text = f"Financial Year: {fy}"
    if from_date and to_date:
        period_text += f"<br/>Period: {from_date} to {to_date}"
    story.append(Paragraph(period_text, styles['PeriodInfo']))
    
    # Summary Box
    summary_data = [
        ['Opening Balance:', f"{opening_balance:,.2f}"],
        ['Total Receipts:', f"{total_receipts:,.2f}"],
        ['Total Payments:', f"{total_payments:,.2f}"],
        ['Net Cash Flow:', f"{total_receipts - total_payments:,.2f}"],
        ['Closing Balance:', f"{balance:,.2f}"]
    ]
    
    if cash_negative > 0:
        summary_data.append(['Cash Negative:', f"{cash_negative:,.2f}"])
    
    summary_table = Table(summary_data, colWidths=[5*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Cash Book Table - Professional Format
    table_data = [
        ['Date', 'Voucher No', 'Type', 'Party Name', 'Narration', 'Mode', 'Cash Received', 'Cash Paid', 'Balance']
    ]
    
    # Add opening balance row
    table_data.append([
        from_date or '01-04-2026',
        '',
        'Opening Balance',
        '',
        '',
        '',
        '',
        '',
        f"{opening_balance:,.2f}"
    ])
    
    # Add transactions
    running_balance = opening_balance
    for entry in cash_entries:
        cash_received = 0.0
        cash_paid = 0.0
        if entry.transaction_type == 'Receipt':
            cash_received = float(entry.amount)
        else:  # Payment
            cash_paid = float(entry.amount)
        
        running_balance += cash_received - cash_paid
        
        table_data.append([
            entry.transaction_date.strftime('%d-%m-%Y'),
            entry.voucher_no,
            entry.transaction_type,
            (entry.party_name or '')[:20],  # Limit to 20 chars
            (entry.narration or '')[:30],   # Limit to 30 chars
            entry.payment_mode or '',
            f"{cash_received:,.2f}" if cash_received > 0 else '',
            f"{cash_paid:,.2f}" if cash_paid > 0 else '',
            f"{running_balance:,.2f}"
        ])
    
    # Add totals row
    total_received = sum(float(entry.amount) for entry in cash_entries if entry.transaction_type == 'Receipt')
    total_paid = sum(float(entry.amount) for entry in cash_entries if entry.transaction_type == 'Payment')
    
    table_data.append([
        '',
        '',
        'TOTALS',
        '',
        '',
        '',
        f"{total_received:,.2f}",
        f"{total_paid:,.2f}",
        f"{balance:,.2f}"
    ])
    
    # Create professional table
    cash_table = Table(table_data, colWidths=[2*cm, 2.5*cm, 1.5*cm, 3*cm, 4*cm, 1.5*cm, 2*cm, 2*cm, 2*cm])
    
    # Professional table styling with soft colors
    cash_table.setStyle(TableStyle([
        # Header styling with soft colors
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#E8F4FD')),  # Soft blue
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#2C3E50')),  # Dark blue-gray
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Opening balance row styling with soft colors
        ('BACKGROUND', (0, 1), (-1, 1), HexColor('#F8F9FA')),  # Very light gray
        ('TEXTCOLOR', (0, 1), (-1, 1), HexColor('#495057')),  # Medium gray
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
        
        # Data rows styling
        ('FONTNAME', (0, 2), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 2), (-1, -2), 8),
        ('ALIGN', (0, 2), (5, -2), 'LEFT'),
        ('ALIGN', (6, 2), (8, -2), 'RIGHT'),
        ('VALIGN', (0, 2), (-1, -2), 'MIDDLE'),
        
        # Totals row styling with soft colors
        ('BACKGROUND', (0, -1), (-1, -1), HexColor('#D4EDDA')),  # Soft green
        ('TEXTCOLOR', (0, -1), (-1, -1), HexColor('#155724')),  # Dark green
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('ALIGN', (0, -1), (5, -1), 'CENTER'),
        ('ALIGN', (6, -1), (8, -1), 'RIGHT'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        
        # Proper borders and spacing
        ('GRID', (0, 0), (-1, -1), 1, HexColor('#DEE2E6')),  # Light gray borders
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        
        # Alternating row colors with soft colors
        ('ROWBACKGROUNDS', (0, 2), (-1, -2), [colors.white, HexColor('#F8F9FA')]),
    ]))
    
    story.append(cash_table)
    
    # Footer
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')} | Page 1", 
        styles['PeriodInfo']
    ))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF value from buffer
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    
    # Create response
    response = Response(
        pdf_data,
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=Cash_Book_Statement.pdf'}
    )
    return response

def export_cash_book_excel(cash_entries, total_receipts, total_payments, balance,
                             opening_balance, cash_negative, from_date, to_date, fy):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from flask import Response
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Book Statement"
    
    # Headers - Professional Cash Book Format
    headers = ['Date', 'Voucher No', 'Type', 'Party Name', 'Narration', 'Payment Mode', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    
    # Opening Balance Row
    ws.cell(row=2, column=1, value=from_date or '01-04-2026')
    ws.cell(row=2, column=3, value='Opening Balance')
    ws.cell(row=2, column=9, value=opening_balance)
    
    # Transactions with Running Balance
    running_balance = opening_balance
    for i, entry in enumerate(cash_entries, 3):
        debit = 0.0
        credit = 0.0
        if entry.transaction_type == 'Payment':
            debit = float(entry.amount)
        else:  # Receipt
            credit = float(entry.amount)
        
        running_balance += credit - debit
        
        ws.cell(row=i, column=1, value=entry.transaction_date.strftime('%d-%m-%Y'))
        ws.cell(row=i, column=2, value=entry.voucher_no)
        ws.cell(row=i, column=3, value=entry.transaction_type)
        ws.cell(row=i, column=4, value=entry.party_name or '')
        ws.cell(row=i, column=5, value=entry.narration)
        ws.cell(row=i, column=6, value=entry.payment_mode)
        ws.cell(row=i, column=7, value=debit if debit > 0 else '')
        ws.cell(row=i, column=8, value=credit if credit > 0 else '')
        ws.cell(row=i, column=9, value=running_balance)
    
    # Totals Row
    total_debits = sum(float(entry.amount) for entry in cash_entries if entry.transaction_type == 'Payment')
    total_credits = sum(float(entry.amount) for entry in cash_entries if entry.transaction_type == 'Receipt')
    
    totals_row = len(cash_entries) + 3
    ws.cell(row=totals_row, column=3, value='Totals')
    ws.cell(row=totals_row, column=7, value=total_debits)
    ws.cell(row=totals_row, column=8, value=total_credits)
    ws.cell(row=totals_row, column=9, value=running_balance)
    
    # Format totals row
    for col in range(1, 10):
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    
    # Format currency columns
    for row in range(2, totals_row + 1):
        for col in [7, 8, 9]:  # Debit, Credit, Balance columns
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = '₹#,##0.00'
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=Cash_Book.xlsx'}
    )
    return response

@cash_book_bp.route("/cash_book/share/<int:entry_id>")
@login_required
def share_entry(entry_id):
    cid = session.get("company_id")
    
    # Get the cash entry
    entry = CashBook.query.filter_by(id=entry_id, company_id=cid).first_or_404()
    
    # Get client mobile number
    mobile_number = None
    if entry.account_id:
        account = Account.query.get(entry.account_id)
        if account and hasattr(account, 'phone') and account.phone:
            mobile_number = account.phone
    
    if not mobile_number:
        flash('No mobile number found for this entry', 'warning')
        return redirect(url_for('cash_book.index'))
    
    # Create WhatsApp share message
    message = f"""
📊 *Cash Book Entry* 📊
Company: SHIVA ICE & MILK PRODUCTS
Date: {entry.transaction_date.strftime('%d-%m-%Y')}
Voucher No: {entry.voucher_no}
Type: {entry.transaction_type}
Amount: ₹{entry.amount:.2f}
Party: {entry.party_name}
Narration: {entry.narration}
Payment Mode: {entry.payment_mode}
    """
    
    # Create WhatsApp URL
    whatsapp_url = f"https://wa.me/?phone=91{mobile_number}&text={message}"
    
    return redirect(whatsapp_url)

@cash_book_bp.route("/cash_book/share-pdf/<int:entry_id>")
@login_required
def share_entry_pdf(entry_id):
    cid = session.get("company_id")
    
    # Get the cash entry
    entry = CashBook.query.filter_by(id=entry_id, company_id=cid).first_or_404()
    
    # Get client mobile number
    mobile_number = None
    if entry.account_id:
        account = Account.query.get(entry.account_id)
        if account and hasattr(account, 'phone') and account.phone:
            mobile_number = account.phone
    
    if not mobile_number:
        flash('No mobile number found for this entry', 'warning')
        return redirect(url_for('cash_book.index'))
    
    # Generate single entry PDF
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    
    doc = SimpleDocTemplate("buffer", pagesize=A4)
    styles = getSampleStyleSheet()
    
    story = []
    
    # Title
    styles.add(ParagraphStyle('CustomTitle', fontSize=14, textColor=colors.black, alignment=1))
    story.append(Paragraph("Cash Book Entry", styles['CustomTitle']))
    story.append(Spacer(1, 12))
    
    # Entry details
    story.append(Paragraph(f"Voucher No: {entry.voucher_no}", styles['Normal']))
    story.append(Paragraph(f"Date: {entry.transaction_date.strftime('%d-%m-%Y')}", styles['Normal']))
    story.append(Paragraph(f"Type: {entry.transaction_type}", styles['Normal']))
    story.append(Paragraph(f"Amount: ₹{entry.amount:.2f}", styles['Normal']))
    story.append(Paragraph(f"Party: {entry.party_name}", styles['Normal']))
    story.append(Paragraph(f"Narration: {entry.narration}", styles['Normal']))
    story.append(Paragraph(f"Payment Mode: {entry.payment_mode}", styles['Normal']))
    story.append(Spacer(1, 12))
    
    doc.build(story)
    
    # Save to BytesIO
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    
    # Create WhatsApp URL with PDF (base64 encoded would be complex, so we'll redirect to a service)
    # For now, we'll just return the PDF and let user share manually
    response = Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=Cash_Entry_{entry.voucher_no}.pdf'}
    )
    
    return response

@cash_book_bp.route("/cash-book/add", methods=["GET", "POST"])
@login_required
def add():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    if request.method == "POST":
        transaction_type = request.form.get("transaction_type", "Receipt")
        
        # Handle amount field safely
        amount_field = request.form.get("amount") or request.form.get("debit_amount") or request.form.get("credit_amount")
        if not amount_field:
            flash("Amount is required!", "danger")
            return render_template("cash_book/form.html", 
                                 accounts=accounts,
                                 next_voucher=next_voucher_no(cid, fy),
                                 today=datetime.now().date().isoformat())
        
        amount = float(amount_field)
        transaction_date = datetime.strptime(request.form.get("transaction_date", datetime.now().date().isoformat()), "%Y-%m-%d").date()
        narration = request.form.get("narration", "").strip()
        party_name = request.form.get("party_name", "").strip()
        payment_mode = request.form.get("payment_mode", "Cash")
        reference_no = request.form.get("reference_no", "").strip()
        account_id = request.form.get("account_id")
        
        # Generate voucher number
        voucher_no = next_voucher_no(cid, fy)
        
        # Create cash book entry
        entry = CashBook(
            company_id=cid,
            fin_year=fy,
            voucher_no=voucher_no,
            transaction_date=transaction_date,
            transaction_type=transaction_type,
            amount=amount,
            narration=narration,
            party_name=party_name,
            payment_mode=payment_mode,
            reference_no=reference_no,
            account_id=int(account_id) if account_id else None
        )
        db.session.add(entry)
        
        # Auto-post journal entry (double-entry accounting)
        if account_id:
            # Find or create Cash Account named "Cash in Hand"
            cash_account = Account.query.filter_by(company_id=cid, is_active=True).filter(
                Account.name == "Cash in Hand"
            ).first()
            if not cash_account:
                cash_account = Account(
                    company_id=cid, name="Cash in Hand",
                    opening_dr=0, opening_cr=0, is_active=True
                )
                db.session.add(cash_account)
                db.session.flush()
            
            journal_voucher_no = next_journal_voucher_no(cid, fy)
            jh = JournalHeader(
                company_id=cid, fin_year=fy,
                voucher_no=journal_voucher_no,
                voucher_type="Cash",
                voucher_date=transaction_date,
                narration=f"Cash Book: {narration}",
                created_by=current_user.id,
            )
            db.session.add(jh)
            db.session.flush()
            
            if transaction_type == "Receipt":
                db.session.add(JournalLine(journal_header_id=jh.id, account_id=cash_account.id, debit=amount, credit=0))
                db.session.add(JournalLine(journal_header_id=jh.id, account_id=int(account_id), debit=0, credit=amount))
            else:
                db.session.add(JournalLine(journal_header_id=jh.id, account_id=int(account_id), debit=amount, credit=0))
                db.session.add(JournalLine(journal_header_id=jh.id, account_id=cash_account.id, debit=0, credit=amount))
        
        db.session.commit()
        
        flash(f"Cash book entry {voucher_no} posted to ledger!", "success")
        return redirect(url_for("cash_book.index"))
    
    # Generate next voucher number for display
    next_voucher = next_voucher_no(cid, fy)
    return render_template("cash_book/form.html", 
                         entry=None,
                         next_voucher=next_voucher,
                         today=date.today().isoformat(),
                         accounts=accounts)

@cash_book_bp.route("/cash-book/edit/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit(entry_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    entry = CashBook.query.filter_by(id=entry_id, company_id=cid, fin_year=fy).first_or_404()
    
    if request.method == "POST":
        entry.transaction_type = request.form.get("transaction_type", "Receipt")
        
        # Handle amount field - check for different possible field names
        amount_field = request.form.get("amount") or request.form.get("debit_amount") or request.form.get("credit_amount")
        if amount_field:
            entry.amount = float(amount_field)
        else:
            flash("Amount is required!", "danger")
            return render_template("cash_book/form.html", 
                                 entry=entry,
                                 next_voucher=entry.voucher_no,
                                 today=entry.transaction_date.isoformat(),
                                 accounts=Account.query.filter_by(company_id=cid).all())
        
        entry.transaction_date = datetime.strptime(request.form.get("transaction_date", entry.transaction_date.isoformat()), "%Y-%m-%d").date()
        entry.narration = request.form.get("narration", "").strip()
        entry.party_name = request.form.get("party_name", "").strip()
        entry.payment_mode = request.form.get("payment_mode", "Cash")
        entry.reference_no = request.form.get("reference_no", "").strip()
        
        # Handle account_id safely
        account_id = request.form.get("account_id")
        if account_id:
            entry.account_id = account_id
        
        db.session.commit()
        flash(f"Cash book entry {entry.voucher_no} updated successfully!", "success")
        return redirect(url_for("cash_book.index"))
    
    return render_template("cash_book/form.html", 
                         entry=entry,
                         next_voucher=entry.voucher_no,
                         today=entry.transaction_date.isoformat(),
                         accounts=accounts)

@cash_book_bp.route("/cash-book/delete/<int:entry_id>", methods=["POST"])
@login_required
def delete(entry_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    entry = CashBook.query.filter_by(id=entry_id, company_id=cid, fin_year=fy).first_or_404()
    voucher_no = entry.voucher_no
    
    db.session.delete(entry)
    db.session.commit()
    
    flash(f"Cash book entry {voucher_no} deleted successfully!", "success")
    return redirect(url_for("cash_book.index"))

@cash_book_bp.route("/cash-book/view/<int:entry_id>")
@login_required
def view(entry_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    entry = CashBook.query.filter_by(id=entry_id, company_id=cid, fin_year=fy).first_or_404()
    
    return render_template("cash_book/view.html", entry=entry)

@cash_book_bp.route("/cash-book/quick-entry", methods=["GET", "POST"])
@login_required
def quick_entry():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    # Get accounts for dropdown
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    # Get parties for dropdown (treat as ledgers)
    parties = Party.query.filter_by(company_id=cid, is_active=True).order_by(Party.name).all()
    
    # Get default cash account (first account with 'Cash' in name or ID 1)
    default_cash = Account.query.filter_by(company_id=cid, is_active=True).filter(
        Account.name.ilike('%cash%')
    ).first()
    
    # Calculate previous cash balance
    previous_balance = 0.0
    if default_cash:
        # Get all cash book entries for this cash account up to today
        today = date.today()
        cash_entries = CashBook.query.filter_by(company_id=cid).filter(
            CashBook.transaction_date <= today
        ).all()
        
        # Calculate balance: Receipts - Payments
        for entry in cash_entries:
            if entry.transaction_type == "Receipt":
                previous_balance += float(entry.amount)
            else:  # Payment
                previous_balance -= float(entry.amount)
    
    if request.method == "POST":
        cash_account_id = request.form.get("cash_account_id")
        payment_mode = request.form.get("payment_mode", "Cash")
        reference_no = request.form.get("reference_no", "").strip()
        
        # Check if this is a single row save or batch save
        single_save = request.form.get("single_save") == "true"
        
        if not cash_account_id:
            flash("Please select a cash account", "danger")
            return redirect(url_for("cash_book.quick_entry"))
        
        # Get cash account
        cash_account = Account.query.get(cash_account_id)
        if not cash_account:
            flash("Invalid cash account selected", "danger")
            return redirect(url_for("cash_book.quick_entry"))
        
        if single_save:
            # Single row save - get first row data only
            account_id_str = request.form.get("account_id")
            cash_received = request.form.get("cash_received")
            cash_paid = request.form.get("cash_paid")
            narration = request.form.get("narration")
            transaction_date = datetime.strptime(request.form.get("row_date"), "%Y-%m-%d").date()
            
            # Update session with last used date
            session["last_txn_date"] = transaction_date.isoformat()
            
            # Parse account_id - can be "acc_123" or "party_123"
            party_name = ""
            actual_account_id = None
            
            if account_id_str:
                if account_id_str.startswith("acc_"):
                    actual_account_id = int(account_id_str.split("_")[1])
                    account = Account.query.get(actual_account_id)
                    party_name = account.name if account else ""
                elif account_id_str.startswith("party_"):
                    party_id = int(account_id_str.split("_")[1])
                    party = Party.query.get(party_id)
                    if party:
                        party_name = party.name
                        # Find or create a ledger account for this party
                        party_account = Account.query.filter_by(company_id=cid, name=party.name).first()
                        if not party_account:
                            party_account = Account(
                                company_id=cid,
                                name=party.name,
                                group_id=21 if transaction_type == "Receipt" else 20,
                                is_active=True
                            )
                            db.session.add(party_account)
                            db.session.flush()
                        actual_account_id = party_account.id
                else:
                    # Fallback for old format
                    actual_account_id = int(account_id_str)
                    account = Account.query.get(actual_account_id)
                    party_name = account.name if account else ""
            
            # Get amount based on transaction type
            amount = 0
            transaction_type = ""
            if cash_received and float(cash_received) > 0:
                amount = float(cash_received)
                transaction_type = "Receipt"
            elif cash_paid and float(cash_paid) > 0:
                amount = float(cash_paid)
                transaction_type = "Payment"
            
            if actual_account_id and transaction_type and amount > 0 and narration:
                # Generate voucher numbers
                cash_voucher_no = next_voucher_no(cid, fy)
                journal_voucher_no = next_journal_voucher_no(cid, fy)
                
                # 1. Create Cash Book Entry
                cash_entry = CashBook(
                    company_id=cid,
                    fin_year=fy,
                    voucher_no=cash_voucher_no,
                    transaction_date=transaction_date,
                    transaction_type=transaction_type,
                    amount=amount,
                    narration=narration,
                    party_name=party_name,
                    payment_mode=payment_mode,
                    reference_no=reference_no,
                    account_id=actual_account_id
                )
                db.session.add(cash_entry)
                
                # 2. Create Journal Entry
                jh = JournalHeader(
                    company_id=cid,
                    fin_year=fy,
                    voucher_no=journal_voucher_no,
                    voucher_type="Auto Cash Entry",
                    voucher_date=transaction_date,
                    narration=f"Auto posted from Cash Book: {narration}",
                    created_by=current_user.id,
                )
                db.session.add(jh)
                db.session.flush()
                
                # 3. Create Journal Lines
                if transaction_type == "Receipt":
                    # Cash Account (Debit), Other Account (Credit)
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=cash_account.id,
                        debit=amount,
                        credit=0,
                    ))
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=actual_account_id,
                        debit=0,
                        credit=amount,
                    ))
                else:  # Payment
                    # Other Account (Debit), Cash Account (Credit)
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=actual_account_id,
                        debit=amount,
                        credit=0,
                    ))
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=cash_account.id,
                        debit=0,
                        credit=amount,
                    ))
                
                # 4. Create Bank Entry if payment mode is bank-related
                if payment_mode in ["Bank", "Online", "NEFT", "RTGS", "IMPS", "UPI"]:
                    bank_accounts = BankAccount.query.filter_by(company_id=cid, is_active=True).all()
                    if bank_accounts:
                        bank_account = bank_accounts[0]
                        
                        bank_voucher_no = f"BNK-{fy}-{JournalHeader.query.filter_by(company_id=cid).count() + 1:04d}"
                        
                        bank_jh = JournalHeader(
                            company_id=cid,
                            fin_year=fy,
                            voucher_no=bank_voucher_no,
                            voucher_type="Bank",
                            voucher_date=transaction_date,
                            narration=f"Auto posted from Cash Book: {narration}",
                            created_by=current_user.id,
                        )
                        db.session.add(bank_jh)
                        db.session.flush()
                        
                        if transaction_type == "Receipt":
                            db.session.add(JournalLine(
                                journal_header_id=bank_jh.id,
                                account_id=bank_account.account_id,
                                debit=amount,
                                credit=0,
                            ))
                            db.session.add(JournalLine(
                                journal_header_id=bank_jh.id,
                                account_id=int(account_id),
                                debit=0,
                                credit=amount,
                            ))
                        else:  # Payment
                            db.session.add(JournalLine(
                                journal_header_id=bank_jh.id,
                                account_id=int(account_id),
                                debit=amount,
                                credit=0,
                            ))
                            db.session.add(JournalLine(
                                journal_header_id=bank_jh.id,
                                account_id=bank_account.account_id,
                                debit=0,
                                credit=amount,
                            ))
                
                db.session.commit()
                flash(f"Cash entry auto-posted to Journal, Ledger, and Banking!", "success")
            else:
                flash("Please fill all required fields", "warning")
        else:
            # Batch save - only save filled rows
            account_ids = request.form.getlist("account_id[]")
            cash_received = request.form.getlist("cash_received[]")
            cash_paid = request.form.getlist("cash_paid[]")
            narrations = request.form.getlist("narration[]")
            row_dates = request.form.getlist("row_date[]")
            
            # Get bank accounts for auto-posting
            bank_accounts = BankAccount.query.filter_by(company_id=cid, is_active=True).all()
            bank_account = bank_accounts[0] if bank_accounts else None
            
            entries_created = 0
            journal_entries_created = 0
            bank_entries_created = 0
            
            for i in range(len(account_ids)):
                # Skip empty rows
                if not account_ids[i] or not narrations[i]:
                    continue
                
                # Check if this row has any amount
                received_amount = float(cash_received[i] or 0)
                paid_amount = float(cash_paid[i] or 0)
                
                if received_amount <= 0 and paid_amount <= 0:
                    continue  # Skip rows with no amounts
                
                # Determine transaction type and amount
                if received_amount > 0:
                    transaction_type = "Receipt"
                    amount = received_amount
                else:
                    transaction_type = "Payment"
                    amount = paid_amount
                
                # Use row-specific date
                if i < len(row_dates) and row_dates[i]:
                    entry_date = datetime.strptime(row_dates[i], "%Y-%m-%d").date()
                    # Update session with the last used row date
                    session["last_txn_date"] = entry_date.isoformat()
                else:
                    # Fallback to session date or today
                    entry_date = datetime.strptime(session.get("last_txn_date") or date.today().isoformat(), "%Y-%m-%d").date()
                
                # Generate voucher numbers
                cash_voucher_no = next_voucher_no(cid, fy)
                journal_voucher_no = next_journal_voucher_no(cid, fy)
                bank_voucher_no = f"BNK-{fy}-{JournalHeader.query.filter_by(company_id=cid).count() + journal_entries_created + 1:04d}"
                
                # Parse account_id - can be "acc_123" or "party_123"
                account_id_str = account_ids[i]
                party_name = ""
                actual_account_id = None
                is_party = False
                
                if account_id_str.startswith("acc_"):
                    actual_account_id = int(account_id_str.split("_")[1])
                    account = Account.query.get(actual_account_id)
                    party_name = account.name if account else ""
                elif account_id_str.startswith("party_"):
                    party_id = int(account_id_str.split("_")[1])
                    party = Party.query.get(party_id)
                    if party:
                        party_name = party.name
                        is_party = True
                        # Find or create a ledger account for this party
                        party_account = Account.query.filter_by(company_id=cid, name=party.name).first()
                        if not party_account:
                            # Create a sundry debtor/creditor account for the party
                            party_account = Account(
                                company_id=cid,
                                name=party.name,
                                group_id=21 if transaction_type == "Receipt" else 20,  # 21=Sundry Debtors, 20=Sundry Creditors
                                is_active=True
                            )
                            db.session.add(party_account)
                            db.session.flush()
                        actual_account_id = party_account.id
                else:
                    # Fallback for old format (just numeric ID)
                    actual_account_id = int(account_id_str)
                    account = Account.query.get(actual_account_id)
                    party_name = account.name if account else ""
                
                # 1. Create Cash Book Entry
                cash_entry = CashBook(
                    company_id=cid,
                    fin_year=fy,
                    voucher_no=cash_voucher_no,
                    transaction_date=entry_date,
                    transaction_type=transaction_type,
                    amount=amount,
                    narration=narrations[i],
                    party_name=party_name,
                    payment_mode=payment_mode,
                    reference_no=reference_no,
                    account_id=actual_account_id
                )
                db.session.add(cash_entry)
                entries_created += 1
                
                # 2. Create Journal Entry
                jh = JournalHeader(
                    company_id=cid,
                    fin_year=fy,
                    voucher_no=journal_voucher_no,
                    voucher_type="Auto Cash Entry",
                    voucher_date=entry_date,
                    narration=f"Auto posted from Cash Book: {narrations[i]}",
                    created_by=current_user.id,
                )
                db.session.add(jh)
                db.session.flush()
                
                # 3. Create Journal Lines
                if transaction_type == "Receipt":
                    # Cash Account (Debit), Other Account (Credit)
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=cash_account.id,
                        debit=amount,
                        credit=0,
                    ))
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=actual_account_id,
                        debit=0,
                        credit=amount,
                    ))
                else:  # Payment
                    # Other Account (Debit), Cash Account (Credit)
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=actual_account_id,
                        debit=amount,
                        credit=0,
                    ))
                    db.session.add(JournalLine(
                        journal_header_id=jh.id,
                        account_id=cash_account.id,
                        debit=0,
                        credit=amount,
                    ))
                
                journal_entries_created += 1
                
                # 4. Create Bank Entry if payment mode is bank-related
                if payment_mode in ["Bank", "Online", "NEFT", "RTGS", "IMPS", "UPI"] and bank_account:
                    bank_jh = JournalHeader(
                        company_id=cid,
                        fin_year=fy,
                        voucher_no=bank_voucher_no,
                        voucher_type="Bank",
                        voucher_date=entry_date,
                        narration=f"Auto posted from Cash Book: {narrations[i]}",
                        created_by=current_user.id,
                    )
                    db.session.add(bank_jh)
                    db.session.flush()
                    
                    if transaction_type == "Receipt":
                        # Bank Account (Debit), Other Account (Credit)
                        db.session.add(JournalLine(
                            journal_header_id=bank_jh.id,
                            account_id=bank_account.account_id,
                            debit=amount,
                            credit=0,
                        ))
                        db.session.add(JournalLine(
                            journal_header_id=bank_jh.id,
                            account_id=actual_account_id,
                            debit=0,
                            credit=amount,
                        ))
                    else:  # Payment
                        # Other Account (Debit), Bank Account (Credit)
                        db.session.add(JournalLine(
                            journal_header_id=bank_jh.id,
                            account_id=actual_account_id,
                            debit=amount,
                            credit=0,
                        ))
                        db.session.add(JournalLine(
                            journal_header_id=bank_jh.id,
                            account_id=bank_account.account_id,
                            debit=0,
                            credit=amount,
                        ))
                    
                    bank_entries_created += 1
            
            if entries_created > 0:
                db.session.commit()
                flash(f"Auto-posted {entries_created} entries to Cash Book, {journal_entries_created} to Journal, and {bank_entries_created} to Banking!", "success")
            else:
                flash("No valid entries to save. Please fill at least one complete row.", "warning")
        
        return redirect(url_for("cash_book.index"))
    
    # Default date to last used or today
    default_date = session.get("last_txn_date") or date.today().isoformat()
    
    return render_template("cash_book/quick_entry.html", 
                         accounts=accounts,
                         parties=parties,
                         default_cash=default_cash,
                         today=default_date,
                         previous_balance=previous_balance)

@cash_book_bp.route("/cash-book/quick-ledger/<int:account_id>")
@login_required
def quick_ledger(account_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    start, end = fy_dates(fy)
    
    # Get account details
    account = Account.query.filter_by(id=account_id, company_id=cid).first_or_404()
    
    # Get opening balance
    opening_balance = 0.0
    if hasattr(account, 'opening_dr') and account.opening_dr:
        opening_balance += float(account.opening_dr)
    if hasattr(account, 'opening_cr') and account.opening_cr:
        opening_balance -= float(account.opening_cr)
    
    # Get all cash book transactions for this account
    cash_entries = CashBook.query.filter(
        CashBook.company_id == cid,
        CashBook.account_id == account_id,
        CashBook.transaction_date >= start,
        CashBook.transaction_date <= end
    ).order_by(CashBook.transaction_date, CashBook.id).all()
    
    # Calculate running balance
    transactions = []
    running_balance = opening_balance
    
    for entry in cash_entries:
        amount = float(entry.amount)
        if entry.transaction_type == "Receipt":
            running_balance += amount
        else:  # Payment
            running_balance -= amount
        
        transactions.append({
            'date': entry.transaction_date,
            'voucher_no': entry.voucher_no,
            'voucher_type': 'Cash',
            'narration': entry.narration,
            'debit': amount if entry.transaction_type == 'Payment' else 0,
            'credit': amount if entry.transaction_type == 'Receipt' else 0,
            'balance': running_balance
        })
    
    return render_template("reports/account_ledger.html",
        account=account,
        opening_balance=opening_balance,
        transactions=transactions,
        closing_balance=running_balance,
        fy=fy)

@cash_book_bp.route("/cash-book/quick-ledger/<int:account_id>/pdf")
@login_required
def quick_ledger_pdf(account_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    start, end = fy_dates(fy)
    
    # Get account details and transactions (reuse logic from quick_ledger)
    account = Account.query.filter_by(id=account_id, company_id=cid).first_or_404()
    
    opening_balance = 0.0
    if hasattr(account, 'opening_dr') and account.opening_dr:
        opening_balance += float(account.opening_dr)
    if hasattr(account, 'opening_cr') and account.opening_cr:
        opening_balance -= float(account.opening_cr)
    
    cash_entries = CashBook.query.filter(
        CashBook.company_id == cid,
        CashBook.account_id == account_id,
        CashBook.transaction_date >= start,
        CashBook.transaction_date <= end
    ).order_by(CashBook.transaction_date, CashBook.id).all()
    
    transactions = []
    running_balance = opening_balance
    
    for entry in cash_entries:
        amount = float(entry.amount)
        if entry.transaction_type == "Receipt":
            running_balance += amount
        else:  # Payment
            running_balance -= amount
        
        transactions.append({
            'date': entry.transaction_date,
            'voucher_no': entry.voucher_no,
            'voucher_type': 'Cash',
            'narration': entry.narration,
            'debit': amount if entry.transaction_type == 'Payment' else 0,
            'credit': amount if entry.transaction_type == 'Receipt' else 0,
            'balance': running_balance
        })
    
    # Generate PDF
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    doc = SimpleDocTemplate("buffer", pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    
    # Add content
    story = []
    
    # Title
    styles.add(ParagraphStyle('CustomTitle', fontSize=16, textColor=colors.black, alignment=1))
    story.append(Paragraph(f"{account.name} - Cash Ledger", styles['CustomTitle']))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Date', 'Voucher No', 'Type', 'Narration', 'Debit', 'Credit', 'Balance']]
    data.append(['Opening Balance', '', '', '', '', '', f"₹{opening_balance:.2f}"])
    
    for txn in transactions:
        data.append([
            txn.date.strftime('%d-%m-%Y'),
            txn.voucher_no,
            txn.voucher_type,
            txn.narration,
            f"₹{txn.debit:.2f}" if txn.debit > 0 else '',
            f"₹{txn.credit:.2f}" if txn.credit > 0 else '',
            f"₹{txn.balance:.2f}"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), colors.white),
        ('TEXTCOLOR', (0, 0), colors.black),
        ('ALIGN', (0, 0), 'CENTER'),
        ('FONTNAME', (0, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), 8),
        ('GRID', (0, 0), 1, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    
    # Return PDF response
    from io import BytesIO
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    
    response = Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={account.name}_Cash_Ledger.pdf'}
    )
    return response

@cash_book_bp.route("/cash-book/quick-ledger/<int:account_id>/excel")
@login_required
def quick_ledger_excel(account_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    start, end = fy_dates(fy)
    
    # Get account details and transactions (reuse logic from quick_ledger)
    account = Account.query.filter_by(id=account_id, company_id=cid).first_or_404()
    
    opening_balance = 0.0
    if hasattr(account, 'opening_dr') and account.opening_dr:
        opening_balance += float(account.opening_dr)
    if hasattr(account, 'opening_cr') and account.opening_cr:
        opening_balance -= float(account.opening_cr)
    
    cash_entries = CashBook.query.filter(
        CashBook.company_id == cid,
        CashBook.account_id == account_id,
        CashBook.transaction_date >= start,
        CashBook.transaction_date <= end
    ).order_by(CashBook.transaction_date, CashBook.id).all()
    
    transactions = []
    running_balance = opening_balance
    
    for entry in cash_entries:
        amount = float(entry.amount)
        if entry.transaction_type == "Receipt":
            running_balance += amount
        else:  # Payment
            running_balance -= amount
        
        transactions.append({
            'date': entry.transaction_date,
            'voucher_no': entry.voucher_no,
            'voucher_type': 'Cash',
            'narration': entry.narration,
            'debit': amount if entry.transaction_type == 'Payment' else 0,
            'credit': amount if entry.transaction_type == 'Receipt' else 0,
            'balance': running_balance
        })
    
    # Generate Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{account.name} - Cash Ledger"
    
    # Headers
    headers = ['Date', 'Voucher No', 'Type', 'Narration', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    
    # Opening Balance
    ws.cell(row=2, column=1, value="Opening Balance")
    ws.cell(row=2, column=7, value=opening_balance)
    
    # Transactions
    for i, txn in enumerate(transactions, 3):
        ws.cell(row=i, column=1, value=txn.date.strftime('%d-%m-%Y'))
        ws.cell(row=i, column=2, value=txn.voucher_no)
        ws.cell(row=i, column=3, value=txn.voucher_type)
        ws.cell(row=i, column=4, value=txn.narration)
        ws.cell(row=i, column=5, value=txn.debit if txn.debit > 0 else 0)
        ws.cell(row=i, column=6, value=txn.credit if txn.credit > 0 else 0)
        ws.cell(row=i, column=7, value=txn.balance)
    
    # Totals
    total_row = len(transactions) + 3
    ws.cell(row=total_row, column=4, value="Totals")
    ws.cell(row=total_row, column=5, value=sum(t['debit'] for t in transactions))
    ws.cell(row=total_row, column=6, value=sum(t['credit'] for t in transactions))
    ws.cell(row=total_row, column=7, value=running_balance)
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={account.name}_Cash_Ledger.xlsx'}
    )
    return response
