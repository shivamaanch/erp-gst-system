from flask import Blueprint, render_template, request, session, flash, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from extensions import db
from models import BankAccount, BankTransaction, BankImportLog, Party, Account, JournalHeader, JournalLine
from datetime import date, datetime
import hashlib, io, re, uuid

banking_bp = Blueprint("banking", __name__)

CASH_KEYWORDS   = ["CASH","CASH DEP","CASH WDL","ATM WDL","ATM CASH","BY CASH","TO CASH"]
SUSPENSE_KWORDS = ["UPI/","IMPS/","NEFT/","RTGS/","CHQ","CHEQUE","ECS","ACH"]
UPI_PATTERN     = re.compile(r"UPI[/-]", re.IGNORECASE)
CHQ_PATTERN     = re.compile(r"(CHQ|CHEQUE|CQ)[^\d]*(\d+)", re.IGNORECASE)

def detect_mode(desc):
    d = (desc or "").upper()
    if any(k in d for k in ["UPI","PHONEPE","GPAY","PAYTM","BHIM"]): return "UPI"
    if any(k in d for k in ["NEFT","RTGS"]): return d[:4]
    if any(k in d for k in ["IMPS"]): return "IMPS"
    if any(k in d for k in ["CHQ","CHEQUE","CQ"]): return "CHQ"
    if any(k in d for k in ["ECS","ACH","NACH"]): return "ECS"
    if any(k in d for k in ["ATM","CASH"]): return "CASH"
    if any(k in d for k in ["IFSC","TRANSFER"]): return "NEFT"
    return "OTHER"

def detect_ledger(desc, mode):
    d = (desc or "").upper()
    if mode == "CASH" or any(k in d for k in CASH_KEYWORDS): return "CASH"
    return "SUSPENSE"

def make_hash(bank_id, txn_date, debit, credit, ref_no, desc):
    raw = f"{bank_id}|{txn_date}|{debit}|{credit}|{ref_no}|{desc}"
    return hashlib.sha256(raw.encode()).hexdigest()

@banking_bp.route("/banking/accounts")
@login_required
def accounts():
    cid = session.get("company_id")
    banks = BankAccount.query.filter_by(company_id=cid, is_active=True).all()
    return render_template("banking/accounts.html", banks=banks)

@banking_bp.route("/banking/accounts/add", methods=["GET","POST"])
@login_required
def add_account():
    cid = session.get("company_id")
    if request.method == "POST":
        acc_no = request.form["account_no"].strip()
        acc_name = request.form["account_name"].strip()
        existing_no = BankAccount.query.filter_by(company_id=cid, account_no=acc_no).first()
        if existing_no:
            flash(f"❌ Account No {acc_no} already exists as '{existing_no.account_name}'", "danger")
            return redirect(url_for("banking.add_account"))
        existing_name = BankAccount.query.filter(
            BankAccount.company_id==cid,
            db.func.lower(BankAccount.account_name)==acc_name.lower()
        ).first()
        if existing_name:
            flash(f"❌ Account name '{acc_name}' already exists!", "danger")
            return redirect(url_for("banking.add_account"))
        b = BankAccount(
            company_id=cid,
            account_name=acc_name,
            bank_name=request.form["bank_name"].strip(),
            account_no=acc_no,
            ifsc=request.form.get("ifsc","").strip().upper(),
            branch=request.form.get("branch","").strip(),
            account_type=request.form.get("account_type","Current"),
            opening_balance=float(request.form.get("opening_balance") or 0)
        )
        db.session.add(b)
        db.session.flush()  # Get the bank account ID
        
        # Create corresponding ledger account for the bank
        from models import Account
        bank_account = Account(
            company_id=cid,
            name=f"{acc_name} - {request.form['bank_name'].strip()}",
            group_name="Bank Accounts",
            opening_dr=float(request.form.get("opening_balance") or 0),
            opening_cr=0,
            is_active=True
        )
        db.session.add(bank_account)
        db.session.flush()  # Get the ledger account ID
        
        # Link bank account to ledger account
        b.account_id = bank_account.id
        
        db.session.commit()
        
        flash(f"✅ Bank account '{acc_name}' created with ledger account!", "success")
        return redirect(url_for("banking.accounts"))
    return render_template("banking/account_form.html")

@banking_bp.route("/banking/template")
@login_required
def download_template():
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bank Statement"
        headers = ["Date(DD/MM/YYYY)","Value Date","Description","Ref No / Cheque No","Debit","Credit","Balance"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a6b3a")
            ws.column_dimensions[cell.column_letter].width = 22
        ws.append(["23/03/2026","23/03/2026","UPI/123456/RAMESH DAIRY","UPI123456",0,5000,25000])
        ws.append(["23/03/2026","23/03/2026","CASH DEPOSIT","",10000,0,15000])
        ws.append(["23/03/2026","23/03/2026","CHQ NO 001234 SURESH KUMAR","001234",8000,0,7000])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, download_name="bank_import_template.xlsx",
                        as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Install openpyxl: pip install openpyxl | Error: {e}","danger")
        return redirect(url_for("banking.accounts"))

@banking_bp.route("/banking/import/<int:bank_id>", methods=["GET","POST"])
@login_required
def import_txns(bank_id):
    cid = session.get("company_id"); fy = session.get("fin_year")
    bank = BankAccount.query.filter_by(id=bank_id, company_id=cid).first_or_404()
    if request.method == "POST":
        try:
            import openpyxl
            f = request.files["file"]
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            batch = str(uuid.uuid4())[:8]
            total = imported = dupes = errors = 0
            notes = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                total += 1
                try:
                    raw_date, val_date, desc, ref_no, debit, credit, balance = (row + (None,)*7)[:7]
                    if not raw_date: continue
                    if isinstance(raw_date, datetime): txn_date = raw_date.date()
                    elif isinstance(raw_date, date): txn_date = raw_date
                    else:
                        for fmt in ["%d/%m/%Y","%d-%m-%Y","%Y-%m-%d","%d/%m/%y"]:
                            try: txn_date = datetime.strptime(str(raw_date).strip(), fmt).date(); break
                            except: continue
                    debit  = float(debit  or 0)
                    credit = float(credit or 0)
                    balance= float(balance or 0)
                    ref_no = str(ref_no or "").strip()
                    desc   = str(desc or "").strip()
                    h = make_hash(bank_id, txn_date, debit, credit, ref_no, desc)
                    if BankTransaction.query.filter_by(hash_key=h).first():
                        dupes += 1; notes.append(f"DUP: {txn_date} {desc[:30]}"); continue
                    mode   = detect_mode(desc)
                    ledger = detect_ledger(desc, mode)
                    t = BankTransaction(
                        company_id=cid, fin_year=fy, bank_account_id=bank_id,
                        txn_date=txn_date,
                        value_date=val_date.date() if isinstance(val_date,datetime) else txn_date,
                        description=desc, ref_no=ref_no,
                        debit=debit, credit=credit, balance=balance,
                        txn_mode=mode, ledger_type=ledger,
                        import_batch=batch, hash_key=h
                    )
                    db.session.add(t); imported += 1
                except Exception as e:
                    errors += 1; notes.append(f"ERR row {total}: {e}")
            db.session.commit()
            log = BankImportLog(company_id=cid, bank_account_id=bank_id,
                file_name=f.filename, total_rows=total, imported=imported,
                duplicates=dupes, errors=errors, notes="\n".join(notes))
            db.session.add(log); db.session.commit()
            flash(f"✅ Imported {imported} | Duplicates skipped: {dupes} | Errors: {errors}", "success")
            return redirect(url_for("banking.transactions", bank_id=bank_id))
        except Exception as e:
            flash(f"❌ Import failed: {e}", "danger")
    return render_template("banking/import.html", bank=bank)

@banking_bp.route("/banking/transactions/<int:bank_id>")
@login_required
def transactions(bank_id):
    cid = session.get("company_id"); fy = session.get("fin_year")
    bank = BankAccount.query.filter_by(id=bank_id, company_id=cid).first_or_404()
    txns = BankTransaction.query.filter_by(
        bank_account_id=bank_id, company_id=cid, fin_year=fy
    ).order_by(BankTransaction.txn_date.desc()).limit(500).all()
    total_dr = sum(float(t.debit)  for t in txns)
    total_cr = sum(float(t.credit) for t in txns)
    suspense  = [t for t in txns if t.ledger_type=="SUSPENSE"]
    cash_txns = [t for t in txns if t.ledger_type=="CASH"]
    return render_template("banking/transactions.html",
        bank=bank, txns=txns, total_dr=total_dr, total_cr=total_cr,
        suspense_count=len(suspense), cash_count=len(cash_txns))

@banking_bp.route("/banking/manual-entry/<int:bank_id>", methods=["GET","POST"])
@login_required
def manual_entry(bank_id):
    cid = session.get("company_id")
    fy = session.get("fin_year")
    bank = BankAccount.query.filter_by(id=bank_id, company_id=cid).first_or_404()
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    if request.method == "POST":
        desc = request.form["description"].strip()
        ref_no = request.form.get("ref_no","").strip()
        debit = float(request.form.get("debit") or 0)
        credit = float(request.form.get("credit") or 0)
        account_id = request.form.get("account_id")
        txn_date = datetime.strptime(request.form["txn_date"],"%Y-%m-%d").date()
        
        if (debit > 0 or credit > 0) and account_id:
            # Create journal entry for double-entry
            jh = JournalHeader(
                company_id=cid,
                fin_year=fy,
                voucher_date=txn_date,
                voucher_type="Bank",
                narration=f"Bank Transaction - {desc}" + (f" (Ref: {ref_no})" if ref_no else "")
            )
            db.session.add(jh)
            db.session.flush()
            
            # Create journal lines
            lines = []
            
            # Bank account line
            if debit > 0:
                # Bank is receiving money (debit)
                lines.append(JournalLine(
                    journal_header_id=jh.id,
                    account_id=bank.account_id or bank_id,  # Use linked account or bank_id
                    debit=debit,
                    credit=0,
                    narration=desc
                ))
                # Other account is giving money (credit)
                lines.append(JournalLine(
                    journal_header_id=jh.id,
                    account_id=int(account_id),
                    debit=0,
                    credit=debit,
                    narration=desc
                ))
            else:
                # Bank is paying money (credit)
                lines.append(JournalLine(
                    journal_header_id=jh.id,
                    account_id=bank.account_id or bank_id,  # Use linked account or bank_id
                    debit=0,
                    credit=credit,
                    narration=desc
                ))
                # Other account is receiving money (debit)
                lines.append(JournalLine(
                    journal_header_id=jh.id,
                    account_id=int(account_id),
                    debit=credit,
                    credit=0,
                    narration=desc
                ))
            
            db.session.add_all(lines)
            db.session.commit()
            
            flash("✅ Bank entry saved successfully!", "success")
            return redirect(url_for("banking.transactions", bank_id=bank_id))
        else:
            flash("❌ Please fill all required fields!", "danger")
    
    return render_template("banking/manual_entry.html", bank=bank, accounts=accounts, today=date.today().isoformat())

@banking_bp.route("/banking/quick-entry", methods=["GET","POST"])
@login_required
def quick_entry():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    # Ensure: every bank ledger account has a corresponding BankAccount row
    bank_ledger_accounts = Account.query.filter(
        Account.company_id == cid,
        Account.is_active == True,
        db.or_(
            Account.group_name.ilike('%bank%'),
            Account.name.ilike('%bank%'),
            Account.name.ilike('%hdfc%'),
            Account.name.ilike('%sbi%'),
            Account.name.ilike('%icici%'),
            Account.name.ilike('%axis%')
        )
    ).order_by(Account.name).all()

    for acc in bank_ledger_accounts:
        existing = BankAccount.query.filter_by(company_id=cid, account_id=acc.id).first()
        if not existing:
            db.session.add(BankAccount(
                company_id=cid,
                account_name=acc.name,
                bank_name=(acc.group_name or "Bank"),
                account_no=None,
                ifsc=None,
                branch=None,
                account_type="Current",
                opening_balance=0,
                is_active=True,
                account_id=acc.id,
            ))
    db.session.commit()

    banks = BankAccount.query.filter_by(company_id=cid, is_active=True).order_by(BankAccount.account_name).all()
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    # Get parties for dropdown (treat as ledgers)
    parties = Party.query.filter_by(company_id=cid, is_active=True).order_by(Party.name).all()
    
    # Calculate previous bank balances
    bank_balances = {}
    today = date.today()
    from utils.filters import fy_dates
    start, end = fy_dates(fy)
    if today < end:
        end = today

    for bank in banks:
        ledger_account_id = bank.account_id
        if not ledger_account_id:
            continue

        journal_lines = JournalLine.query.join(JournalHeader).filter(
            JournalHeader.company_id == cid,
            JournalHeader.fin_year == fy,
            JournalHeader.voucher_date >= start,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False,
            JournalLine.account_id == ledger_account_id,
        ).all()

        balance = float(bank.opening_balance or 0)
        for line in journal_lines:
            balance += float(line.debit or 0)
            balance -= float(line.credit or 0)

        bank_balances[bank.id] = balance
    
    if request.method == "POST":
        bank_account_id = request.form.get("bank_account_id")
        transaction_type = request.form.get("transaction_type", "Deposit")
        payment_mode = request.form.get("payment_mode", "Cash")
        reference_no = request.form.get("reference_no", "").strip()
        party_name = request.form.get("party_name", "").strip()
        
        # Check if this is a single row save or batch save
        single_save = request.form.get("single_save") == "true"
        
        bank_obj = BankAccount.query.filter_by(id=bank_account_id, company_id=cid, is_active=True).first()
        if not bank_obj or not bank_obj.account_id:
            flash("❌ Please select a valid bank account", "danger")
            return redirect(url_for("banking.quick_entry"))

        def _auto_narration(is_receipt: bool) -> str:
            mode = payment_mode or "Bank"
            ref = reference_no.strip()
            party = party_name.strip()
            if is_receipt:
                base = f"Amount received in bank via {mode}" + (f" ({ref})" if ref else "")
            else:
                base = f"Payment made from bank via {mode}" + (f" ({ref})" if ref else "")
            if party:
                return f"{base} - {party}"
            return base

        def _post_one(entry_date, other_account_id, dr_amount, cr_amount, narration_text):
            # 1) Banking transaction (for Banking -> Transactions screen)
            bt = BankTransaction(
                company_id=cid,
                fin_year=fy,
                bank_account_id=bank_obj.id,
                txn_date=entry_date,
                description=narration_text,
                ref_no=reference_no,
                debit=dr_amount,
                credit=cr_amount,
                txn_mode=(payment_mode or "Bank"),
                ledger_type="BANK",
            )
            db.session.add(bt)

            # 2) Journal entry (for Ledger / BS / P&L)
            jh = JournalHeader(
                company_id=cid,
                fin_year=fy,
                voucher_no=None,
                voucher_type="Bank",
                voucher_date=entry_date,
                narration=narration_text,
                created_by=current_user.id,
            )
            db.session.add(jh)
            db.session.flush()

            if dr_amount > 0:
                # Receipt: Bank Dr, Other Cr
                db.session.add(JournalLine(
                    journal_header_id=jh.id,
                    account_id=bank_obj.account_id,
                    debit=dr_amount,
                    credit=0,
                    narration=narration_text,
                ))
                db.session.add(JournalLine(
                    journal_header_id=jh.id,
                    account_id=int(other_account_id),
                    debit=0,
                    credit=dr_amount,
                    narration=narration_text,
                ))
            else:
                # Payment: Other Dr, Bank Cr
                db.session.add(JournalLine(
                    journal_header_id=jh.id,
                    account_id=int(other_account_id),
                    debit=cr_amount,
                    credit=0,
                    narration=narration_text,
                ))
                db.session.add(JournalLine(
                    journal_header_id=jh.id,
                    account_id=bank_obj.account_id,
                    debit=0,
                    credit=cr_amount,
                    narration=narration_text,
                ))

        if single_save:
            row_date = request.form.get("row_date", date.today().isoformat())
            account_id_str = request.form.get("account_id")
            dr_amount = float(request.form.get("debit") or 0)
            cr_amount = float(request.form.get("credit") or 0)
            narration_text = (request.form.get("narration") or "").strip()
            
            # Parse account_id - can be "acc_123" or "party_123"
            actual_account_id = None
            if account_id_str:
                if account_id_str.startswith("acc_"):
                    actual_account_id = int(account_id_str.split("_")[1])
                elif account_id_str.startswith("party_"):
                    party_id = int(account_id_str.split("_")[1])
                    party = Party.query.get(party_id)
                    if party:
                        # Find or create a ledger account for this party
                        party_account = Account.query.filter_by(company_id=cid, name=party.name).first()
                        if not party_account:
                            party_account = Account(
                                company_id=cid,
                                name=party.name,
                                group_id=21 if dr_amount > 0 else 20,  # 21=Sundry Debtors, 20=Sundry Creditors
                                is_active=True
                            )
                            db.session.add(party_account)
                            db.session.flush()
                        actual_account_id = party_account.id
                else:
                    # Fallback for old format
                    actual_account_id = int(account_id_str)

            if not actual_account_id or (dr_amount <= 0 and cr_amount <= 0):
                flash("❌ Please fill account and amount", "danger")
                return redirect(url_for("banking.quick_entry"))

            entry_date = datetime.strptime(row_date, "%Y-%m-%d").date()
            session["last_txn_date"] = entry_date.isoformat()
            if not narration_text:
                narration_text = _auto_narration(is_receipt=(dr_amount > 0))

            _post_one(entry_date, actual_account_id, dr_amount, cr_amount, narration_text)
            db.session.commit()
            flash("✅ Bank entry saved successfully!", "success")
            return redirect(url_for("banking.quick_entry"))

        # Batch save: save all filled rows (no overall balancing required)
        row_dates = request.form.getlist("row_date[]")
        account_ids = request.form.getlist("account_id[]")
        debits = request.form.getlist("debit[]")
        credits = request.form.getlist("credit[]")
        narrations = request.form.getlist("narration[]")

        created = 0
        for i in range(len(account_ids)):
            if not account_ids[i]:
                continue
            dr_amount = float(debits[i] or 0)
            cr_amount = float(credits[i] or 0)
            if dr_amount <= 0 and cr_amount <= 0:
                continue

            # Parse account_id - can be "acc_123" or "party_123"
            actual_account_id = None
            account_id_str = account_ids[i]
            
            if account_id_str.startswith("acc_"):
                actual_account_id = int(account_id_str.split("_")[1])
            elif account_id_str.startswith("party_"):
                party_id = int(account_id_str.split("_")[1])
                party = Party.query.get(party_id)
                if party:
                    # Find or create a ledger account for this party
                    party_account = Account.query.filter_by(company_id=cid, name=party.name).first()
                    if not party_account:
                        party_account = Account(
                            company_id=cid,
                            name=party.name,
                            group_id=21 if dr_amount > 0 else 20,
                            is_active=True
                        )
                        db.session.add(party_account)
                        db.session.flush()
                    actual_account_id = party_account.id
            else:
                # Fallback for old format
                actual_account_id = int(account_id_str)

            entry_date = date.today()
            if i < len(row_dates) and row_dates[i]:
                entry_date = datetime.strptime(row_dates[i], "%Y-%m-%d").date()
            session["last_txn_date"] = entry_date.isoformat()

            narration_text = (narrations[i] or "").strip()
            if not narration_text:
                narration_text = _auto_narration(is_receipt=(dr_amount > 0))

            _post_one(entry_date, actual_account_id, dr_amount, cr_amount, narration_text)
            created += 1

        if created:
            db.session.commit()
            flash(f"✅ Saved {created} bank entr{'y' if created==1 else 'ies'}", "success")
        else:
            flash("⚠️ No filled rows found to save", "warning")

        return redirect(url_for("banking.quick_entry"))
    
    # Get default bank account
    default_bank = banks[0] if banks else None
    
    # Default date to last used or today
    default_date = session.get("last_txn_date") or date.today().isoformat()
    
    return render_template("banking/quick_entry.html", 
                         banks=banks,
                         accounts=accounts,
                         parties=parties,
                         default_bank=default_bank,
                         today=default_date,
                         bank_balances=bank_balances)
