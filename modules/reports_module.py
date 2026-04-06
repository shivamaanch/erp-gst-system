# modules/reports_module.py
from flask import Blueprint, render_template, request, session, jsonify, redirect, url_for, flash
from flask_login import login_required
from extensions import db
from models import Party, Bill, Company, Account, JournalLine, JournalHeader
from sqlalchemy import func
from datetime import date

reports_bp = Blueprint("reports", __name__)

def get_account_balances():
    """Get individual account balances for Balance Sheet"""
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    # Get financial year dates
    from utils.filters import fy_dates
    start, end = fy_dates(fy)
    
    journal_sums = db.session.query(
        JournalLine.account_id.label("account_id"),
        func.coalesce(func.sum(JournalLine.debit), 0).label("dr"),
        func.coalesce(func.sum(JournalLine.credit), 0).label("cr"),
    ).join(
        JournalHeader, JournalHeader.id == JournalLine.journal_header_id
    ).filter(
        JournalHeader.company_id == cid,
        JournalHeader.fin_year == fy,
        JournalHeader.voucher_date >= start,
        JournalHeader.voucher_date <= end,
        JournalHeader.is_cancelled == False,
    ).group_by(JournalLine.account_id).subquery()

    accounts_with_sums = db.session.query(
        Account,
        journal_sums.c.dr,
        journal_sums.c.cr,
    ).outerjoin(
        journal_sums, journal_sums.c.account_id == Account.id
    ).filter(
        Account.company_id == cid,
        Account.is_active == True,
    ).all()

    balances = {}
    for acc, dr, cr in accounts_with_sums:
        opening = float(acc.opening_dr or 0) - float(acc.opening_cr or 0)
        txn_balance = float(dr or 0) - float(cr or 0)
        balance = opening + txn_balance
        if abs(balance) > 0.01:
            balances[acc.id] = {
                'id': acc.id,
                'name': acc.name,
                'group': acc.group_name,
                'balance': balance,
            }

    return balances

@reports_bp.route("/reports/hub")
@login_required
def hub():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    
    # Get sales and purchase data
    total_sales     = float(db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Sales", is_cancelled=False).scalar() or 0)
    total_purchases = float(db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Purchase", is_cancelled=False).scalar() or 0)
    sales_count     = Bill.query.filter_by(company_id=cid, fin_year=fy, bill_type="Sales", is_cancelled=False).count()
    purchase_count  = Bill.query.filter_by(company_id=cid, fin_year=fy, bill_type="Purchase", is_cancelled=False).count()
    
    # Calculate actual financial position from database
    from utils.filters import fy_dates
    start, end = fy_dates(fy)
    
    # Calculate financial position (Total Assets) from FY-filtered balances
    balances = get_account_balances()
    financial_position = sum(v['balance'] for v in balances.values() if float(v.get('balance') or 0) > 0)
    
    # Calculate net profit
    net_profit = total_sales - total_purchases
    
    return render_template("reports/hub.html", fy=fy,
                           total_sales=total_sales, total_purchases=total_purchases,
                           sales_count=sales_count, purchase_count=purchase_count,
                           net_profit=net_profit,
                           financial_position=financial_position,
                           current_date=date.today())

@reports_bp.route("/reports/quick-ledger")
@login_required
def quick_ledger():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    from utils.filters import fy_dates
    start, end = fy_dates(fy)
    
    # Get all accounts for quick search
    accounts = Account.query.filter_by(company_id=cid, is_active=True).order_by(Account.name).all()
    
    # Pre-compute account balances from journal lines
    account_bal_query = db.session.query(
        JournalLine.account_id,
        func.sum(JournalLine.debit).label("dr"),
        func.sum(JournalLine.credit).label("cr"),
    ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
    ).filter(
        JournalHeader.company_id == cid,
        JournalHeader.voucher_date >= start,
        JournalHeader.voucher_date <= end,
        JournalHeader.is_cancelled == False
    ).group_by(JournalLine.account_id).all()
    
    account_txn_bal = {r.account_id: float(r.dr or 0) - float(r.cr or 0) for r in account_bal_query}
    
    # Get all parties for quick search
    from models import Party, Bill, CashBook
    parties = Party.query.filter_by(company_id=cid, is_active=True).order_by(Party.name).all()
    
    # Pre-compute party balances from bills + cash book
    party_bill_totals = {}
    bill_rows = db.session.query(
        Bill.party_id,
        Bill.bill_type,
        func.sum(Bill.total_amount).label("total")
    ).filter(
        Bill.company_id == cid, Bill.fin_year == fy, Bill.is_cancelled == False
    ).group_by(Bill.party_id, Bill.bill_type).all()
    
    for r in bill_rows:
        if r.party_id not in party_bill_totals:
            party_bill_totals[r.party_id] = 0.0
        amt = float(r.total or 0)
        if r.bill_type == 'Purchase':
            party_bill_totals[r.party_id] += amt  # We owe = debit for party
        elif r.bill_type in ('Sales', 'Sale'):
            party_bill_totals[r.party_id] -= amt  # They owe = credit for party
    
    # Combine accounts and parties - skip accounts that duplicate party names
    all_ledgers = []
    party_names_lower = {p.name.lower() for p in parties}
    
    # Add accounts (skip those that are just party duplicates)
    for account in accounts:
        if account.account_type == 'Party' and account.name.lower() in party_names_lower:
            continue  # Skip - party already represents this entity
        
        opening = float(account.opening_dr or 0) - float(account.opening_cr or 0)
        txn_bal = account_txn_bal.get(account.id, 0.0)
        
        # Check if this is a milk-related account
        is_milk = False
        if account.name:
            name_lower = account.name.lower()
            if any(keyword in name_lower for keyword in ['milk', 'purchase', 'sale']):
                is_milk = True
        if account.group_name and 'milk' in account.group_name.lower():
            is_milk = True
        
        all_ledgers.append({
            'id': account.id,
            'name': account.name,
            'type': 'Account',
            'entity': 'account',
            'group_name': account.group_name or '',
            'account_type': account.account_type or '',
            'balance': round(opening + txn_bal, 2),
            'is_milk': is_milk
        })
    
    # Add parties with actual calculated balance
    for party in parties:
        opening = float(party.opening_balance or 0) * (1 if party.balance_type == 'Dr' else -1)
        bill_bal = party_bill_totals.get(party.id, 0.0)
        
        # Check if this party has milk transactions
        is_milk = False
        if party.name:
            name_lower = party.name.lower()
            if any(keyword in name_lower for keyword in ['milk', 'purchase', 'sale']):
                is_milk = True
        
        all_ledgers.append({
            'id': party.id,
            'name': party.name,
            'type': 'Party',
            'entity': 'party',
            'party_type': party.party_type or '',
            'gstin': party.gstin or '',
            'phone': party.phone or '',
            'balance': round(opening + bill_bal, 2),
            'is_milk': is_milk
        })
    
    # Sort by name
    all_ledgers.sort(key=lambda x: x['name'].lower())
    
    return render_template("reports/quick_ledger.html", accounts=all_ledgers)

@reports_bp.route("/reports/delete-transaction", methods=["POST"])
@login_required
def delete_transaction():
    """Delete a transaction completely from all places"""
    try:
        cid = session.get("company_id")
        fy = session.get("fin_year")
        
        data = request.get_json()
        voucher_no = data.get("voucher_no", "")
        transaction_date = data.get("transaction_date", "")
        entity_id = data.get("entity_id", 0)
        entity_type = data.get("entity_type", "")
        
        if not voucher_no:
            return jsonify({"success": False, "message": "Voucher number is required"})
        
        print(f"DEBUG: Deleting transaction - Voucher: {voucher_no}, Date: {transaction_date}, Entity: {entity_id} ({entity_type})")
        
        # Find the journal header for this voucher
        from models import JournalHeader, JournalLine, Bill, BillItem, MilkTransaction, CashBook
        
        journal_header = JournalHeader.query.filter_by(
            company_id=cid,
            voucher_no=voucher_no
        ).first()
        
        if not journal_header:
            return jsonify({"success": False, "message": "Transaction not found"})
        
        print(f"DEBUG: Found journal header {journal_header.id}")
        
        # Delete all journal lines for this header
        journal_lines = JournalLine.query.filter_by(journal_header_id=journal_header.id).all()
        for line in journal_lines:
            print(f"DEBUG: Deleting journal line {line.id}")
            db.session.delete(line)
        
        # Delete the journal header
        print(f"DEBUG: Deleting journal header {journal_header.id}")
        db.session.delete(journal_header)
        
        # Find and delete any bills linked to this voucher
        bills = Bill.query.filter(
            Bill.company_id == cid,
            Bill.narration.like(f"%{voucher_no}%")
        ).all()
        
        for bill in bills:
            print(f"DEBUG: Deleting bill {bill.id} - {bill.bill_no}")
            # Delete bill items
            BillItem.query.filter_by(bill_id=bill.id).delete()
            # Find and delete milk transactions linked to this bill
            milk_txns = MilkTransaction.query.filter_by(bill_id=bill.id).all()
            for milk_txn in milk_txns:
                print(f"DEBUG: Deleting milk transaction {milk_txn.id}")
                db.session.delete(milk_txn)
            # Delete the bill
            db.session.delete(bill)
        
        # Find and delete milk transactions directly linked to this voucher (if no bill)
        milk_txns = MilkTransaction.query.filter(
            MilkTransaction.company_id == cid,
            MilkTransaction.narration.like(f"%{voucher_no}%")
        ).all()
        
        for milk_txn in milk_txns:
            print(f"DEBUG: Deleting milk transaction {milk_txn.id} (direct)")
            db.session.delete(milk_txn)
        
        # Find and delete cash book entries linked to this voucher
        cash_entries = CashBook.query.filter(
            CashBook.company_id == cid,
            CashBook.voucher_no.like(f"%{voucher_no}%")
        ).all()
        
        for cash_entry in cash_entries:
            print(f"DEBUG: Deleting cash book entry {cash_entry.id}")
            db.session.delete(cash_entry)
        
        # Commit all deletions
        db.session.commit()
        
        print(f"DEBUG: Transaction {voucher_no} deleted successfully from all places")
        
        return jsonify({
            "success": True, 
            "message": f"Transaction {voucher_no} deleted successfully from all records"
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error deleting transaction: {str(e)}")
        return jsonify({
            "success": False, 
            "message": f"Error deleting transaction: {str(e)}"
        })

@reports_bp.route("/reports/account-ledger/<int:account_id>")
@login_required
def account_ledger(account_id):
    from utils.filters import fy_dates
    from models import MilkTransaction
    
    cid = session.get("company_id")
    fy = session.get("fin_year")
    entity = (request.args.get("entity") or "").lower()
    
    # Resolve entity explicitly to avoid account/party id collisions
    account = None
    party = None

    from models import Party
    report_type = request.args.get("report_type", "")
    
    if entity == "party":
        party = Party.query.filter_by(id=account_id, company_id=cid).first()
    elif entity == "account":
        if account_id == 0 and report_type:
            # Special case: sales/purchases from P&L
            account = None
        else:
            account = Account.query.filter_by(id=account_id, company_id=cid).first()
    else:
        account = Account.query.filter_by(id=account_id, company_id=cid).first()
        if not account:
            party = Party.query.filter_by(id=account_id, company_id=cid).first()
    
    # Determine entity details
    if account:
        entity_name = account.name
        entity_group = account.group_name or "Default Group"
        entity_type = "Account"
    elif party:
        entity_name = party.name
        entity_group = f"Party - {party.party_type}"
        entity_type = "Party"
    elif account_id == 0 and report_type:
        # Special case from P&L links
        entity_name = report_type.title()
        entity_group = "P&L Summary"
        entity_type = "Report"
    else:
        entity_name = "Unknown Entity"
        entity_group = "Default Group"
        entity_type = "Unknown"
    
    # Get financial year dates
    start, end = fy_dates(fy)
    
    # Get transactions for this entity
    transactions = []
    
    if account:
        # Get Account transactions from JournalLines
        transactions = db.session.query(
            JournalLine,
            JournalHeader
        ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            JournalLine.account_id == account_id,
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date >= start,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False
        ).order_by(JournalHeader.voucher_date, JournalHeader.id).all()
    
    elif party:
        # Get Party transactions from Bills (milk entries, invoices, etc.)
        from models import Bill
        transactions = Bill.query.filter(
            Bill.party_id == party.id,
            Bill.company_id == cid,
            Bill.fin_year == fy,
            Bill.is_cancelled == False
        ).order_by(Bill.bill_date, Bill.id).all()
    
    elif entity_type == "Report" and report_type:
        # Special case: Sales or Purchases from P&L
        from models import Bill
        if report_type == "sales":
            transactions = Bill.query.filter(
                Bill.company_id == cid,
                Bill.fin_year == fy,
                Bill.bill_type == "Sales",
                Bill.is_cancelled == False
            ).order_by(Bill.bill_date, Bill.id).all()
        elif report_type == "purchases":
            transactions = Bill.query.filter(
                Bill.company_id == cid,
                Bill.fin_year == fy,
                Bill.bill_type == "Purchase",
                Bill.is_cancelled == False
            ).order_by(Bill.bill_date, Bill.id).all()
    
    # Calculate opening balance
    opening_balance = 0.0
    if account:
        opening_balance = float(account.opening_dr or 0) - float(account.opening_cr or 0)
    elif party:
        opening_balance = float(party.opening_balance or 0) * (1 if party.balance_type == 'Dr' else -1)
    
    # Calculate running balance
    running_balance = opening_balance
    transaction_data = []
    
    if account:
        # Process Account transactions (JournalLines)
        for line, header in transactions:
            debit = float(line.debit or 0)
            credit = float(line.credit or 0)
            
            if debit > 0:
                running_balance += debit
            else:
                running_balance -= credit
            
            # Check if this is a milk transaction and get details
            qty_liters = None; fat = None; snf = None; rate = None; shift = None; bf_clr = None
            fat_value = 0.0; snf_value = 0.0
            
            narration = line.narration or header.narration or ''
            if 'milk' in narration.lower():
                # Try to find milk transaction by matching narration pattern
                try:
                    # Extract voucher number from narration if it's a milk transaction
                    import re
                    mlk_match = re.search(r'MLK-(\d+)', narration)
                    if mlk_match:
                        milk_txn_id = int(mlk_match.group(1))
                        milk_txn = MilkTransaction.query.filter_by(id=milk_txn_id, company_id=cid).first()
                        if milk_txn:
                            qty_liters = float(milk_txn.qty_liters) if milk_txn.qty_liters is not None else None
                            fat = float(milk_txn.fat) if milk_txn.fat is not None else None
                            snf = float(milk_txn.snf) if milk_txn.snf is not None else None
                            rate = float(milk_txn.rate) if milk_txn.rate is not None else None
                            shift = milk_txn.shift
                            bf_clr = f"BF:{fat} / CLR:{milk_txn.clr}" if fat is not None and getattr(milk_txn, 'clr', None) is not None else None
                            
                            # Calculate fat and SNF values
                            if qty_liters and (fat or snf):
                                daily_rate = float(rate or 0)
                                if daily_rate > 0:
                                    try:
                                        from modules.milk import compute_component_breakdown
                                        calc = compute_component_breakdown(qty_liters, fat or 0, snf or 0, daily_rate)
                                        fat_value = round(calc["bf_kgs"] * calc["bf_rate"], 2)
                                        snf_value = round(calc["snf_kgs"] * calc["snf_rate"], 2)
                                    except Exception:
                                        fat_value = 0.0
                                        snf_value = 0.0
                except Exception:
                    pass
            
            transaction_data.append({
                'date': header.voucher_date,
                'voucher_no': header.voucher_no,
                'type': header.voucher_type or 'Journal',
                'narration': narration,
                'debit': debit,
                'credit': credit,
                'balance': running_balance,
                'qty_liters': qty_liters,
                'fat_percentage': fat,
                'snf_percentage': snf,
                'rate': rate,
                'basic_amount': debit + credit,
                'fat_value': fat_value,
                'snf_value': snf_value,
                'total_amount': debit + credit,
                'shift': shift,
                'bf_clr': bf_clr
            })
    
    elif party:
        # Process Party transactions from ALL sources
        all_party_txns = []
        
        # A. Bills (milk entries, invoices, etc.)
        for bill in transactions:
            amount = float(bill.total_amount or 0)
            narration = bill.narration or f'{bill.bill_type} - {bill.bill_no}'
            
            qty_liters = None; fat = None; snf = None; rate = None; shift = None; bf_clr = None
            try:
                milk_txn = MilkTransaction.query.filter_by(company_id=cid, bill_id=bill.id).order_by(MilkTransaction.id.desc()).first()
                if milk_txn:
                    qty_liters = float(milk_txn.qty_liters) if milk_txn.qty_liters is not None else None
                    fat = float(milk_txn.fat) if milk_txn.fat is not None else None
                    snf = float(milk_txn.snf) if milk_txn.snf is not None else None
                    rate = float(milk_txn.rate) if milk_txn.rate is not None else None
                    shift = milk_txn.shift
                    bf_clr = f"BF:{fat} / CLR:{milk_txn.clr}" if fat is not None and getattr(milk_txn, 'clr', None) is not None else None
            except Exception:
                pass
            
            if bill.bill_type == 'Purchase':
                debit = amount; credit = 0
                print(f"DEBUG: Purchase Bill {bill.bill_no} - Amount: {amount}, Debit: {debit}, Credit: {credit}")
            elif bill.bill_type in ['Sale', 'Sales']:
                debit = 0; credit = amount
                print(f"DEBUG: Sale Bill {bill.bill_no} - Amount: {amount}, Debit: {debit}, Credit: {credit}")
            else:
                debit = amount; credit = 0
                print(f"DEBUG: Other Bill {bill.bill_no} (Type: {bill.bill_type}) - Amount: {amount}, Debit: {debit}, Credit: {credit}")
            
            # Calculate fat and SNF values for milk transactions using unified milk formula
            fat_value = 0.0
            snf_value = 0.0
            if qty_liters and (fat or snf):
                daily_rate = float(rate or 0)
                if daily_rate > 0:
                    try:
                        from modules.milk import compute_component_breakdown
                        calc = compute_component_breakdown(qty_liters, fat or 0, snf or 0, daily_rate)
                        fat_value = round(calc["bf_kgs"] * calc["bf_rate"], 2)
                        snf_value = round(calc["snf_kgs"] * calc["snf_rate"], 2)
                    except Exception:
                        fat_value = 0.0
                        snf_value = 0.0
            
            all_party_txns.append({
                'date': bill.bill_date, 'voucher_no': bill.bill_no,
                'type': bill.bill_type or 'Bill', 'voucher_type': bill.bill_type or 'Bill',
                'narration': narration, 'debit': debit, 'credit': credit,
                'qty_liters': qty_liters, 'fat_percentage': fat, 'snf_percentage': snf,
                'rate': rate, 'basic_amount': amount, 'fat_value': fat_value, 'snf_value': snf_value,
                'total_amount': amount, 'shift': shift, 'bf_clr': bf_clr
            })
        
        # B. Cash Book entries for this party (by party_name match)
        from models import CashBook
        cash_entries = CashBook.query.filter(
            CashBook.company_id == cid,
            CashBook.fin_year == fy,
            db.func.lower(CashBook.party_name) == party.name.lower()
        ).order_by(CashBook.transaction_date, CashBook.id).all()
        
        for cb in cash_entries:
            amt = float(cb.amount or 0)
            if cb.transaction_type == 'Receipt':
                debit = 0; credit = amt
            else:
                debit = amt; credit = 0
            all_party_txns.append({
                'date': cb.transaction_date, 'voucher_no': cb.voucher_no,
                'type': f'Cash {cb.transaction_type}', 'voucher_type': 'Cash',
                'narration': cb.narration or '', 'debit': debit, 'credit': credit,
                'qty_liters': None, 'fat_percentage': None, 'snf_percentage': None,
                'rate': None, 'basic_amount': amt, 'fat_value': 0.0, 'snf_value': 0.0,
                'total_amount': amt, 'shift': None, 'bf_clr': None
            })
        
        # C. Journal entries for matching account (by name)
        matching_account = Account.query.filter(
            Account.company_id == cid,
            db.func.lower(Account.name) == party.name.lower()
        ).first()
        
        if matching_account:
            jl_entries = db.session.query(JournalLine, JournalHeader).join(
                JournalHeader, JournalHeader.id == JournalLine.journal_header_id
            ).filter(
                JournalLine.account_id == matching_account.id,
                JournalHeader.company_id == cid,
                JournalHeader.voucher_date >= start,
                JournalHeader.voucher_date <= end,
                JournalHeader.is_cancelled == False
            ).order_by(JournalHeader.voucher_date, JournalHeader.id).all()
            
            for line, header in jl_entries:
                debit = float(line.debit or 0)
                credit = float(line.credit or 0)
                all_party_txns.append({
                    'date': header.voucher_date, 'voucher_no': header.voucher_no,
                    'type': header.voucher_type or 'Journal', 'voucher_type': header.voucher_type or 'Journal',
                    'narration': line.narration or header.narration or '',
                    'debit': debit, 'credit': credit,
                    'qty_liters': None, 'fat_percentage': None, 'snf_percentage': None,
                    'rate': None, 'basic_amount': 0, 'fat_value': 0.0, 'snf_value': 0.0,
                    'total_amount': 0, 'shift': None, 'bf_clr': None
                })
        
        # Sort all party transactions by date
        all_party_txns.sort(key=lambda x: x['date'])
        
        # Calculate running balance
        for txn in all_party_txns:
            running_balance += float(txn['debit'] or 0) - float(txn['credit'] or 0)
            txn['balance'] = running_balance
            transaction_data.append(txn)
    
    # Calculate totals
    total_debits = sum(t['debit'] for t in transaction_data)
    total_credits = sum(t['credit'] for t in transaction_data)
    closing_balance = running_balance
    net_change = closing_balance - opening_balance
    total_transactions = len(transaction_data)
    
    return render_template("reports/account_ledger.html",
                         account=account,
                         party=party,
                         entity_name=entity_name,
                         entity_group=entity_group,
                         entity_type=entity_type,
                         transactions=transaction_data,
                         fy=fy,
                         opening_balance=opening_balance,
                         closing_balance=closing_balance,
                         net_change=net_change,
                         total_transactions=total_transactions,
                         total_debits=total_debits,
                         total_credits=total_credits)

@reports_bp.route("/reports/account-ledger/<int:account_id>/pdf")
@login_required
def account_ledger_pdf(account_id):
    from utils.filters import fy_dates
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm
    from io import BytesIO
    from flask import Response
    
    cid = session.get("company_id")
    fy = session.get("fin_year")
    company = Company.query.get(cid)
    entity = request.args.get("entity", "account")
    
    # Get account/party details
    account_name = "Unknown"
    if entity == "party":
        p = Party.query.get(account_id)
        account_name = p.name if p else "Unknown Party"
    else:
        acc = Account.query.get(account_id) if account_id > 0 else None
        account_name = acc.name if acc else "Default Account"
    
    start, end = fy_dates(fy)
    
    # Get transactions
    transactions = []
    if entity == "account" and account_id > 0:
        transactions = db.session.query(JournalLine, JournalHeader).join(
            JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            JournalLine.account_id == account_id,
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date >= start,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False
        ).order_by(JournalHeader.voucher_date, JournalLine.id).all()
    
    # Build PDF into BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=1*cm, bottomMargin=1*cm,
                            leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('CompanyName', fontSize=16, textColor=colors.HexColor('#1F4E79'),
                              alignment=1, spaceAfter=4, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle('SubHeader', fontSize=10, textColor=colors.black, alignment=1, spaceAfter=8))
    
    story = []
    
    # Professional company header
    company_name = company.name if company else 'Company'
    story.append(Paragraph(company_name, styles['CompanyName']))
    if company and company.address:
        story.append(Paragraph(company.address, styles['SubHeader']))
    story.append(Paragraph(f"Account Ledger: {account_name} | FY: {fy}", styles['SubHeader']))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Date', 'Voucher No', 'Type', 'Narration', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)']]
    col_widths = [2*cm, 3*cm, 2.5*cm, 9*cm, 3*cm, 3*cm, 3*cm]
    
    running_balance = 0.0
    for line, header in transactions:
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)
        running_balance += debit - credit
        data.append([
            header.voucher_date.strftime('%d-%m-%Y'),
            header.voucher_no,
            header.voucher_type or 'Journal',
            (line.narration or header.narration or '')[:60],
            f"{debit:,.2f}" if debit > 0 else '',
            f"{credit:,.2f}" if credit > 0 else '',
            f"{running_balance:,.2f}"
        ])
    
    # Totals row
    tot_dr = sum(float(l.debit or 0) for l, h in transactions)
    tot_cr = sum(float(l.credit or 0) for l, h in transactions)
    data.append(['', '', '', 'TOTAL', f"{tot_dr:,.2f}", f"{tot_cr:,.2f}", f"{running_balance:,.2f}"])
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (2, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D6E4F0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Generated on {date.today().strftime('%d-%b-%Y')} | {company_name}", styles['SubHeader']))
    
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={account_name}_Ledger.pdf'}
    )

@reports_bp.route("/reports/account-ledger/<int:account_id>/excel")
@login_required
def account_ledger_excel(account_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from flask import Response
    from utils.filters import fy_dates
    
    cid = session.get("company_id")
    fy = session.get("fin_year")
    company = Company.query.get(cid)
    entity = request.args.get("entity", "account")
    
    # Get name
    if entity == "party":
        p = Party.query.get(account_id)
        account_name = p.name if p else "Unknown Party"
    else:
        acc = Account.query.get(account_id) if account_id > 0 else None
        account_name = acc.name if acc else "Default Account"
    
    start, end = fy_dates(fy)
    
    transactions = []
    if entity == "account" and account_id > 0:
        transactions = db.session.query(JournalLine, JournalHeader).join(
            JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            JournalLine.account_id == account_id,
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date >= start,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False
        ).order_by(JournalHeader.voucher_date, JournalLine.id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    
    # Company header row
    company_name = company.name if company else 'Company'
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value=company_name)
    c.font = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    c = ws.cell(row=2, column=1, value=f"Account Ledger: {account_name} | FY: {fy}")
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="center")
    
    # Column headers at row 4
    headers = ['Date', 'Voucher No', 'Type', 'Narration', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)']
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Data rows
    running_balance = 0.0
    row = 5
    for line, header in transactions:
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)
        running_balance += debit - credit
        
        vals = [
            header.voucher_date.strftime('%d-%m-%Y'),
            header.voucher_no,
            header.voucher_type or 'Journal',
            line.narration or header.narration or '',
            debit if debit > 0 else '',
            credit if credit > 0 else '',
            running_balance
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col >= 5 and isinstance(val, (int, float)) and val != '':
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
        row += 1
    
    # Totals row
    tot_dr = sum(float(l.debit or 0) for l, h in transactions)
    tot_cr = sum(float(l.credit or 0) for l, h in transactions)
    for col, val in enumerate(['', '', '', 'TOTAL', tot_dr, tot_cr, running_balance], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin_border
        if col >= 5 and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={account_name}_Ledger.xlsx'}
    )

@reports_bp.route("/reports/profit_loss")
@login_required
def profit_loss():
    cid  = session.get("company_id")
    fy   = session.get("fin_year")
    from utils.filters import fy_dates
    start, end = fy_dates(fy)
    
    # Get company
    company = Company.query.get(cid)
    
    # 1. Sales and Purchases from Bills (milk entries + invoices)
    sales_from_bills = float(db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Sales", is_cancelled=False).scalar() or 0)
    purchase_from_bills = float(db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Purchase", is_cancelled=False).scalar() or 0)
    
    # 2. Income/Expense from Journal entries (cash book auto-posts, manual journals)
    INCOME_GROUPS = ["Sales", "Direct Income", "Indirect Income", "Other Income"]
    EXPENSE_GROUPS = ["Purchase", "Direct Expense", "Indirect Expense", "Depreciation",
                      "Cost of Goods Sold", "Salary", "Rent", "Utilities"]
    
    def journal_group_totals(groups):
        rows = db.session.query(
            Account.group_name,
            func.sum(JournalLine.debit).label("dr"),
            func.sum(JournalLine.credit).label("cr"),
        ).join(JournalLine, JournalLine.account_id == Account.id
        ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            Account.company_id == cid,
            Account.group_name.in_(groups),
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date.between(start, end),
            JournalHeader.is_cancelled == False,
        ).group_by(Account.group_name).all()
        return {r.group_name: float(r.cr or 0) - float(r.dr or 0) for r in rows}
    
    income_from_journals = journal_group_totals(INCOME_GROUPS)
    expense_from_journals = journal_group_totals(EXPENSE_GROUPS)
    
    journal_income_total = sum(income_from_journals.values())
    journal_expense_total = sum(expense_from_journals.values())
    
    # Combine: Bills + Journal entries
    sales_total = sales_from_bills + max(journal_income_total, 0)
    purchase_total = purchase_from_bills + abs(min(journal_expense_total, 0))
    
    # Calculate profits
    gross_profit = sales_total - purchase_total
    net_profit = gross_profit
    total_expenses = purchase_total
    total_income = sales_total
    
    # Use horizontal template by default
    return render_template("reports/profit_loss_horizontal.html", 
                           fy=fy,
                           company=company,
                           sales=sales_total,
                           purchases=purchase_total,
                           gross_profit=gross_profit,
                           net_profit=net_profit,
                           total_expenses=total_expenses,
                           total_income=total_income,
                           income_details=income_from_journals,
                           expense_details=expense_from_journals,
                           fiscal_year_end=f"31 Mar {int(fy[:4])+1}",
                           current_date=date.today())

@reports_bp.route("/reports/balance_sheet")
@login_required
def balance_sheet():
    cid  = session.get("company_id")
    fy   = session.get("fin_year")
    
    # Get company
    company = Company.query.get(cid)
    
    # Get account balances for detailed display
    account_balances = get_account_balances()
    
    # Get Fixed Assets data from Fixed Assets Schedule
    from modules.fixed_assets import get_fixed_assets_schedule
    fixed_assets_schedule = get_fixed_assets_schedule(cid, fy)
    
    # Calculate simplified totals (keeping existing logic for compatibility)
    creditors = Party.query.filter_by(company_id=cid, party_type="Creditor", is_active=True).all()
    debtors   = Party.query.filter_by(company_id=cid, party_type="Debtor", is_active=True).all()
    creditors = sum(float(c.opening_balance or 0) for c in creditors if (c.balance_type or "Cr") == "Cr")
    debtors   = sum(float(d.opening_balance or 0) for d in debtors if (d.balance_type or "Dr") == "Dr")
    
    # Capital and reserves (simplified)
    share_capital_acc = Account.query.filter_by(company_id=cid, name="Share Capital", is_active=True).first()
    capital = float(share_capital_acc.opening_dr or 0) if share_capital_acc else 0
    
    reserves_acc = Account.query.filter_by(company_id=cid, name="General Reserve", is_active=True).first()
    retained = float(reserves_acc.opening_dr or 0) if reserves_acc else 0
    
    # Calculate total fixed assets from schedule
    total_fixed_assets = fixed_assets_schedule['total_closing_wdv'] if fixed_assets_schedule else 0.0
    
    total_liabilities = capital + retained + creditors
    total_assets      = debtors + total_fixed_assets + (total_liabilities - debtors - total_fixed_assets)  # balance
    
    # Use horizontal template by default
    return render_template("reports/balance_sheet_horizontal.html", 
                           fy=fy,
                           company=company,
                           debtors=debtors, 
                           creditors=creditors,
                           retained_earnings=retained, 
                           capital=capital,
                           total_liabilities=total_liabilities, 
                           total_assets=total_assets,
                           account_balances=account_balances,
                           fixed_assets_schedule=fixed_assets_schedule,
                           fiscal_year_end=f"31 Mar {int(fy[:4])+1}",
                           current_date=date.today())

@reports_bp.route("/reports/trial_balance")
@login_required
def trial_balance():
    cid  = session.get("company_id")
    fy   = session.get("fin_year")
    parties = Party.query.filter_by(company_id=cid, is_active=True).all()
    tb_data = []
    total_debit = total_credit = 0
    for p in parties:
        bal = float(p.opening_balance or 0)
        if bal == 0: continue
        if (p.balance_type or "Dr") == "Dr":
            tb_data.append({"account": p.name, "debit": bal, "credit": 0})
            total_debit += bal
        else:
            tb_data.append({"account": p.name, "debit": 0, "credit": bal})
            total_credit += bal
    return render_template("reports/trial_balance.html", fy=fy,
                           tb_data=tb_data, total_debit=total_debit, total_credit=total_credit,
                           fiscal_year_end=f"31 Mar {int(fy[:4])+1}",
                           current_date=date.today())

@reports_bp.route("/reports/ledger")
@reports_bp.route("/reports/ledger/<int:party_id>")
@login_required
def ledger(party_id=None):
    cid     = session.get("company_id")
    fy      = session.get("fin_year")
    parties = Party.query.filter_by(company_id=cid, is_active=True).order_by(Party.name).all()
    if party_id is None and parties:
        party_id = parties[0].id
    party   = Party.query.filter_by(id=party_id, company_id=cid).first() if party_id else None
    ledger_entries = []
    if party:
        bills    = Bill.query.filter_by(company_id=cid, party_id=party_id, fin_year=fy,
                                        is_cancelled=False).order_by(Bill.bill_date).all()
        running  = float(party.opening_balance or 0) if (party.balance_type or "Dr")=="Dr"                    else -float(party.opening_balance or 0)
        for b in bills:
            amt = float(b.total_amount)
            if b.bill_type == "Sales":
                running += amt
                ledger_entries.append({"date": b.bill_date, "particulars": f"Sales Invoice {b.bill_no}",
                                       "debit": amt, "credit": 0, "balance": running})
            else:
                running -= amt
                ledger_entries.append({"date": b.bill_date, "particulars": f"Purchase Bill {b.bill_no}",
                                       "debit": 0, "credit": amt, "balance": running})
    return render_template("reports/ledger.html", party=party, parties=parties,
                           fy=fy, ledger_entries=ledger_entries)

@reports_bp.route("/reports/outstanding")
@login_required
def outstanding():
    cid     = session.get("company_id")
    fy      = session.get("fin_year")
    debtors = Party.query.filter_by(company_id=cid, party_type="Debtor", is_active=True).all()
    creditors = Party.query.filter_by(company_id=cid, party_type="Creditor", is_active=True).all()
    return render_template("reports/outstanding.html", fy=fy, debtors=debtors, creditors=creditors)

@reports_bp.route("/reports/balance_sheet/pdf")
@login_required
def balance_sheet_pdf():
    """Download Balance Sheet as PDF"""
    from flask import flash, redirect, url_for
    flash("PDF download feature coming soon!", "info")
    return redirect(url_for('reports.balance_sheet'))

@reports_bp.route("/reports/balance_sheet/excel")
@login_required
def balance_sheet_excel():
    """Download Balance Sheet as Excel"""
    from flask import flash, redirect, url_for
    flash("Excel download feature coming soon!", "info")
    return redirect(url_for('reports.balance_sheet'))

@reports_bp.route("/reports/profit_loss/pdf")
@login_required
def profit_loss_pdf():
    """Download P&L as PDF"""
    from flask import flash, redirect, url_for
    flash("PDF download feature coming soon!", "info")
    return redirect(url_for('reports.profit_loss'))

@reports_bp.route("/reports/profit_loss/excel")
@login_required
def profit_loss_excel():
    """Download P&L as Excel"""
    from flask import flash, redirect, url_for
    flash("Excel download feature coming soon!", "info")
    return redirect(url_for('reports.profit_loss'))
