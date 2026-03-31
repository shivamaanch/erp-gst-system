# modules/enhanced_invoice.py
from flask import Blueprint, render_template, request, session, flash, redirect, url_for, send_file, jsonify
from flask_login import login_required
from extensions import db
from models import Bill, Party, Item, Company, BillItem
from datetime import datetime, date
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json

enhanced_invoice_bp = Blueprint("enhanced_invoice", __name__)


def _fy_aliases(fy: str):
    if not fy:
        return []
    aliases = {fy}
    # Accept both 2025-26 and 25-26 style FY strings
    if len(fy) == 7 and fy[4] == "-":
        aliases.add(f"{fy[2:4]}-{fy[5:7]}")
    elif len(fy) == 5 and fy[2] == "-":
        aliases.add(f"20{fy}")
    return list(aliases)

# Dynamic table templates for different business types
INVOICE_TEMPLATES = {
    "Trading": {
        "name": "Standard Trading Invoice",
        "columns": ["Item", "Description", "HSN", "Qty", "Unit", "Rate", "Disc%", "Taxable", "GST%", "GST Amount", "Total"],
        "calculations": ["basic", "gst", "total"]
    },
    "Manufacturing": {
        "name": "Manufacturing Invoice",
        "columns": ["Item", "Description", "HSN", "Qty", "Unit", "Rate", "Raw Material Cost", "Manufacturing Cost", "Disc%", "Taxable", "GST%", "GST Amount", "Total"],
        "calculations": ["manufacturing_cost", "gst", "total"]
    },
    "Service": {
        "name": "Service Invoice",
        "columns": ["Service", "Description", "Hours", "Rate", "Disc%", "Taxable", "GST%", "GST Amount", "Total"],
        "calculations": ["service", "gst", "total"]
    },
    "Milk": {
        "name": "Milk Collection Invoice",
        "columns": ["Farmer", "Date", "Fat %", "SNF %", "Quantity (Ltrs)", "Rate/Ltr", "Basic Amount", "Fat Value", "SNF Value", "Total", "Deductions", "Net Amount"],
        "calculations": ["milk_formula", "deductions", "net"]
    }
}

@enhanced_invoice_bp.route("/enhanced-invoice/create", methods=["GET", "POST"])
@login_required
def create_invoice():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    btype = request.args.get("type", "Sales")
    party_types = ["Debtor", "Both", "Customer"] if btype == "Sales" else ["Creditor", "Supplier", "Both", "Vendor"]
    
    if request.method == "POST":
        # Get company to determine business type
        company = Company.query.get(cid) if cid else None
        template_type = request.form.get("template_type", (company.business_type if company else "Trading") or "Trading")

        is_igst = request.form.get("is_igst") == "on"
        item_ids = request.form.getlist("item_id[]")
        qtys = request.form.getlist("quantity[]")
        rates = request.form.getlist("rate[]")
        gst_rates = request.form.getlist("gst_rate[]")

        if not item_ids:
            flash("Add at least one item", "danger")
            return redirect(url_for("enhanced_invoice.create_invoice", type=btype))
        
        # Create bill
        bill = Bill(
            company_id=cid,
            party_id=int(request.form["party_id"]),
            bill_type=btype,
            bill_no=request.form["bill_no"],
            bill_date=datetime.strptime(request.form["bill_date"], "%Y-%m-%d").date(),
            fin_year=fy,
            narration=request.form.get("narration", ""),
            template_type=template_type,
            created_by=session.get("user_id", 1)
        )
        db.session.add(bill)
        
        total_taxable = 0.0
        total_cgst = 0.0
        total_sgst = 0.0
        total_igst = 0.0

        for i, iid in enumerate(item_ids):
            if not iid:
                continue
            try:
                qty = float(qtys[i]) if i < len(qtys) else 0.0
                rate = float(rates[i]) if i < len(rates) else 0.0
                gst_rate = float(gst_rates[i]) if i < len(gst_rates) and gst_rates[i] != "" else 0.0
            except Exception:
                qty = 0.0
                rate = 0.0
                gst_rate = 0.0

            taxable_amount = round(qty * rate, 2)
            gst_amount = round(taxable_amount * gst_rate / 100, 2)

            if is_igst:
                igst = gst_amount
                cgst = 0.0
                sgst = 0.0
            else:
                cgst = round(gst_amount / 2, 2)
                sgst = round(gst_amount / 2, 2)
                igst = 0.0

            db.session.add(
                BillItem(
                    bill=bill,
                    item_id=int(iid),
                    qty=qty,
                    rate=rate,
                    taxable_amount=taxable_amount,
                    gst_rate=gst_rate,
                    cgst=cgst,
                    sgst=sgst,
                    igst=igst,
                )
            )

            total_taxable += taxable_amount
            total_cgst += cgst
            total_sgst += sgst
            total_igst += igst
        
        # Update bill totals
        bill.taxable_amount = round(total_taxable, 2)
        bill.cgst = round(total_cgst, 2)
        bill.sgst = round(total_sgst, 2)
        bill.igst = round(total_igst, 2)
        bill.total_amount = round(total_taxable + total_cgst + total_sgst + total_igst, 2)

        # Optional TDS/TCS
        try:
            bill.tds_rate = float(request.form.get("tds_rate") or 0)
        except Exception:
            bill.tds_rate = 0
        try:
            bill.tcs_rate = float(request.form.get("tcs_rate") or 0)
        except Exception:
            bill.tcs_rate = 0
        
        db.session.commit()
        session["last_txn_date"] = bill.bill_date.isoformat()
        
        flash(f"{btype} invoice created successfully!", "success")
        return redirect(url_for("enhanced_invoice.list_invoices", type=btype))
    
    # Get data for form
    parties = Party.query.filter_by(company_id=cid, is_active=True).filter(
        Party.party_type.in_(party_types)
    ).order_by(Party.name).all()
    items = Item.query.filter_by(company_id=cid, is_active=True).all()
    from models import MilkRateChart
    charts = MilkRateChart.query.filter_by(company_id=cid).all()
    
    company = Company.query.get(cid)
    
    # Generate bill number
    last_bill = Bill.query.filter_by(company_id=cid, bill_type=btype, fin_year=fy).order_by(Bill.id.desc()).first()
    bill_no = f"{btype[0]}{fy[:4]}{(last_bill.id if last_bill else 0) + 1:04d}"
    
    default_bill_date = session.get("last_txn_date") or date.today().isoformat()

    return render_template("enhanced_invoice/create.html",
                         btype=btype, bill_no=bill_no,
                         parties=parties, items=items,
                         bill_date=default_bill_date,
                         fy=fy,
                         company=company,
                         charts=charts,
                         templates=INVOICE_TEMPLATES)

@enhanced_invoice_bp.route("/enhanced-invoice/delete/<int:bill_id>", methods=["POST"])
@login_required
def delete_invoice(bill_id):
    """Delete an invoice"""
    try:
        cid = session.get("company_id")
        bill = Bill.query.filter_by(id=bill_id, company_id=cid).first()
        
        if not bill:
            return jsonify({"success": False, "message": "Invoice not found"})
        
        if bill.is_cancelled:
            return jsonify({"success": False, "message": "Invoice is already cancelled"})
        
        # Delete bill items first
        BillItem.query.filter_by(bill_id=bill_id).delete()
        
        # Find and delete any milk entry linked to this bill (not just clear bill_id)
        from models import MilkTransaction
        linked_milk = MilkTransaction.query.filter_by(bill_id=bill_id).first()
        if linked_milk:
            print(f"DEBUG: Deleting linked milk entry {linked_milk.id}")
            db.session.delete(linked_milk)
        
        # Delete the bill
        db.session.delete(bill)
        db.session.commit()
        print("DEBUG: Invoice and linked milk entry deleted successfully")
        
        return jsonify({"success": True, "message": "Invoice deleted successfully"})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})

@enhanced_invoice_bp.route("/enhanced-invoice/list")
@login_required
def list_invoices():
    cid = session.get("company_id")
    fy = session.get("fin_year")
    btype = request.args.get("type", "Sales")
    
    fy_list = _fy_aliases(fy)
    q = Bill.query.filter(
        Bill.company_id == cid,
        Bill.bill_type == btype,
        Bill.is_cancelled == False,
    )
    if fy_list:
        q = q.filter(Bill.fin_year.in_(fy_list))
    bills = q.order_by(Bill.bill_date.desc()).all()
    
    return render_template("enhanced_invoice/list.html", bills=bills, btype=btype)

@enhanced_invoice_bp.route("/enhanced-invoice/print/<int:bill_id>")
@login_required
def print_invoice(bill_id):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    company = Company.query.get(cid)
    
    # Get bill items
    items = BillItem.query.filter_by(bill_id=bill_id).all()
    
    return render_template("enhanced_invoice/print_pdf.html",
                         bill=bill, party=party, company=company,
                         items=items)

@enhanced_invoice_bp.route("/enhanced-invoice/export/<int:bill_id>/<format>")
@login_required
def export_invoice(bill_id, format):
    cid = session.get("company_id")
    bill = Bill.query.filter_by(id=bill_id, company_id=cid).first_or_404()
    party = Party.query.get(bill.party_id)
    company = Company.query.get(cid)
    
    if format == "excel":
        return export_invoice_excel(bill, party, company)
    elif format == "pdf":
        return export_invoice_pdf(bill, party, company)
    else:
        flash("Invalid format", "danger")
        return redirect(url_for("enhanced_invoice.print_invoice", bill_id=bill_id))

def export_invoice_excel(bill, party, company):
    """Export invoice to Excel with beautiful formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Excel sheet titles cannot contain certain characters like /, \\, ?, *, [, ] and must be <= 31 chars
    safe_bill_no = (bill.bill_no or "").replace("/", "-").replace("\\", "-").replace("?", "-").replace("*", "-").replace("[", "(").replace("]", ")")
    ws.title = ("Invoice " + safe_bill_no)[:31]
    
    # Company header
    ws.merge_cells("A1:H1")
    ws["A1"] = company.name
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = company.address or ""
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A3:H3")
    ws["A3"] = f"GSTIN: {company.gstin or 'Not Available'} | PAN: {company.pan or 'Not Available'}"
    ws["A3"].alignment = Alignment(horizontal="center")
    
    # Invoice details
    row = 5
    ws[f"A{row}"] = "Invoice Details:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    ws[f"A{row}"] = f"Invoice No: {bill.bill_no}"
    ws[f"C{row}"] = f"Date: {bill.bill_date.strftime('%d-%m-%Y')}"
    row += 1
    
    ws[f"A{row}"] = f"Party: {party.name}"
    ws[f"C{row}"] = f"GSTIN: {party.gstin or 'Not Available'}"
    row += 2
    
    # Items table
    template = INVOICE_TEMPLATES.get(bill.template_type, INVOICE_TEMPLATES["Trading"])
    headers = template["columns"]
    
    ws[f"A{row}"] = "Item Details:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    # Header row
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
    row += 1
    
    # Data rows (sample data - you'll need to fetch actual items)
    if hasattr(bill, 'items') and bill.items:
        for item in bill.items:
            ws.cell(row=row, column=1, value=item.item.name if item.item else "Item")
            ws.cell(row=row, column=2, value=f"{item.qty or 0}")
            ws.cell(row=row, column=3, value=f"₹{item.rate or 0}")
            ws.cell(row=row, column=4, value=f"₹{item.taxable_amount or 0}")
            ws.cell(row=row, column=5, value=f"₹{(item.taxable_amount or 0) * 0.18}")
            ws.cell(row=row, column=6, value=f"₹{item.taxable_amount * 1.18}")
            row += 1
    
    # Totals
    row += 1
    ws[f"A{row}"] = "Summary:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    ws[f"A{row}"] = "Taxable Amount:"
    ws[f"E{row}"] = f"₹{bill.taxable_amount or 0}"
    row += 1
    
    ws[f"A{row}"] = "CGST:"
    ws[f"E{row}"] = f"₹{bill.cgst or 0}"
    row += 1
    
    ws[f"A{row}"] = "SGST:"
    ws[f"E{row}"] = f"₹{bill.sgst or 0}"
    row += 1
    
    ws[f"A{row}"] = "IGST:"
    ws[f"E{row}"] = f"₹{bill.igst or 0}"
    row += 1
    
    ws[f"A{row}"] = "Total Amount:"
    ws[f"E{row}"] = f"₹{bill.total_amount or 0}"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"E{row}"].font = Font(bold=True)
    
    # Bank details
    row += 2
    ws[f"A{row}"] = "Bank Details:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    ws[f"A{row}"] = "Account Name: " + (company.name or "")
    ws[f"A{row+1}"] = "Account Number: " + ("[Your Bank Account]" or "")
    ws[f"A{row+2}"] = "IFSC Code: " + ("[Your IFSC]" or "")
    ws[f"A{row+3}"] = "Bank Name: " + ("[Your Bank Name]" or "")
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name=f"Invoice_{bill.bill_no}.xlsx", as_attachment=True,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def export_invoice_pdf(bill, party, company):
    """Export invoice to PDF (placeholder - you'll need to add PDF library like WeasyPrint)"""
    # For now, return the HTML template which can be printed to PDF
    return render_template("enhanced_invoice/print_pdf.html",
                         bill=bill, party=party, company=company)

@enhanced_invoice_bp.route("/api/template-fields/<template_type>")
@login_required
def get_template_fields(template_type):
    """API to get fields for dynamic template"""
    template = INVOICE_TEMPLATES.get(template_type, INVOICE_TEMPLATES["Trading"])
    return jsonify(template)
